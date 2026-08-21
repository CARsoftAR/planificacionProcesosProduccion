from django.shortcuts import render, get_object_or_404, redirect
from django.db import transaction, connections
from django.http import JsonResponse, HttpResponse
from .gantt_logic import get_gantt_data
from .services import get_planificacion_data, get_all_machines
from itertools import groupby
from operator import itemgetter
from .models import (
    PrioridadManual, MaquinaConfig, HorarioMaquina, 
    TaskDependency, HiddenTask, Scenario, ProyectoPrioridad,
    PlannedTask
)
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json
import re
from django.contrib import messages
from .planning_service import calculate_timeline
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

...

@csrf_exempt
def link_tasks(request):
    """
    API to create a dependency: Successor depends on Predecessor.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        pred_id = body.get('predecessor_id')
        succ_id = body.get('successor_id')
        
        if not pred_id or not succ_id:
            return JsonResponse({'error': 'Missing IDs'}, status=400)
            
        TaskDependency.objects.using('default').get_or_create(
            predecessor_id=pred_id,
            successor_id=succ_id
        )
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def unlink_tasks(request):
    """
    API to remove a dependency.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        body = json.loads(request.body)
        pred_id = body.get('predecessor_id')
        succ_id = body.get('successor_id')
        
        TaskDependency.objects.using('default').filter(
            predecessor_id=pred_id,
            successor_id=succ_id
        ).delete()
        
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_active_scenario(request, scenario_id=None):
    """
    Helper to resolve the active scenario from URL, session, or POST body.
    """
    url_scenario_id = request.GET.get('scenario_id')
    
    # If explicitly empty in URL, we don't wipe the session blindly
    # because it might be a reload from a provisional state.
    if url_scenario_id == "" or url_scenario_id == "null":
        scenario_id = None
    elif url_scenario_id:
        scenario_id = url_scenario_id
    
    # Removed logic that previously wiped the session if plan_mode was 'manual'.
    # This ensures smooth persistence when returning from Gantt.
    if not scenario_id and request.method == 'POST':
        try:
            body = json.loads(request.body)
            scenario_id = body.get('scenario_id')
        except:
            pass
            
    if not scenario_id:
        scenario_id = request.session.get('last_scenario_id')
        
    active_scenario = None
    if scenario_id and str(scenario_id).isdigit():
        active_scenario = Scenario.objects.using('default').filter(id=scenario_id).first()
        
    if not active_scenario:
        # Fallback to Principal (Official)
        active_scenario = Scenario.objects.using('default').filter(es_principal=True).first()
        
    if not active_scenario:
        active_scenario = Scenario.objects.using('default').first()
        
    if not active_scenario:
        try:
            active_scenario = Scenario.objects.using('default').create(
                nombre="Plan Principal (Oficial)",
                es_principal=True,
                proyectos=""
            )
            print("[Scenario] Escenario por defecto creado.")
        except Exception as e:
            print(f"[Scenario ERROR] {e}")
        
    if active_scenario:
        request.session['last_scenario_id'] = str(active_scenario.id)
    else:
        request.session['last_scenario_id'] = None
        
    return active_scenario

@csrf_exempt
def reset_planning(request):
    """
    API to clear manual planning (Visual Priorities, Virtual Moves) for a set of Orders.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        ids = body.get('ids', [])
        active_scenario = get_active_scenario(request)

        if active_scenario:
            print(f"DEBUG: reset_planning called for Scenario: {active_scenario.nombre} (ID: {active_scenario.id}) with {len(ids)} IDs")
        else:
            print(f"DEBUG: reset_planning called for NO SCENARIO with {len(ids)} IDs")
        
        # If no IDs on screen, try to find them by project filter
        proyectos = body.get('proyectos')
        if not ids and proyectos:
             if isinstance(proyectos, str):
                  proj_list = [p.strip() for p in proyectos.split(',') if p.strip()]
             else:
                  proj_list = proyectos
             
             if proj_list:
                  # Use the same data fetching logic to find all OPs for these projects
                  from .services import get_planificacion_data
                  erp_data = get_planificacion_data({'proyectos': proj_list})
                  ids = [int(d['Idorden']) for d in erp_data if d.get('Idorden')]
                  print(f"DEBUG: reset_planning - Expanded to {len(ids)} tasks via Project lookup: {proj_list}")

        if not ids:
             return JsonResponse({'status': 'ignored', 'message': 'No IDs provided and no projects found to reset'})
        
        # 1. Clear Priorities and Virtual Moves FOR THIS SCENARIO ONLY
        deleted_prio, _ = PrioridadManual.objects.using('default').filter(id_orden__in=ids, scenario=active_scenario).delete()
        
        # 2. Clear Dependencies (Manual ones)
        # These are currently global
        deleted_dep_pred, _ = TaskDependency.objects.using('default').filter(predecessor_id__in=ids).delete()
        deleted_dep_succ, _ = TaskDependency.objects.using('default').filter(successor_id__in=ids).delete()
        
        # 3. Clear HIDDEN Status FOR THIS SCENARIO ONLY
        deleted_hidden, _ = HiddenTask.objects.using('default').filter(id_orden__in=ids, scenario=active_scenario).delete()
        
        return JsonResponse({'status': 'ok', 'message': f'Reset {len(ids)} tasks'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def commit_gantt_snapshot(request):
    """
    API que congela el estado actual de las planillas como Snapshot para el Gantt.
    
    Lógica:
    1. Resetea enviado_a_gantt=False en TODAS las tareas del escenario (limpia el snapshot viejo).
    2. Excluye las tareas ocultas por el usuario (HiddenTask).
    3. Marca enviado_a_gantt=True en las tareas visibles restantes.
    4. Devuelve los project codes congelados para que el frontend navegue al Gantt.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        active_scenario = get_active_scenario(request)
        
        # Validar si hay tareas en la planificación actual en lugar de requerir un escenario
        qs_planned = PlannedTask.objects.using('default')
        if active_scenario:
            qs_planned = qs_planned.filter(scenario=active_scenario)
        else:
            qs_planned = qs_planned.filter(scenario__isnull=True)
            
        if not qs_planned.exists():
            return JsonResponse({'error': 'No hay tareas en la planificación actual'}, status=400)
        
        # Paso 1: Resetear el snapshot anterior completo
        qs_planned.update(enviado_a_gantt=False)
        
        # Paso 2: Obtener IDs de tareas ocultas en este escenario
        qs_hidden = HiddenTask.objects.using('default')
        if active_scenario:
            qs_hidden = qs_hidden.filter(scenario=active_scenario)
        else:
            qs_hidden = qs_hidden.filter(scenario__isnull=True)
            
        hidden_ids = list(qs_hidden.values_list('id_orden', flat=True))
        
        # Paso 3: Marcar como "en el Gantt" solo las tareas visibles
        updated = qs_planned.exclude(id_orden__in=hidden_ids).update(enviado_a_gantt=True)
        
        # Paso 4: Leer los proyectos congelados para devolver al frontend
        qs_frozen = PlannedTask.objects.using('default').filter(enviado_a_gantt=True)
        if active_scenario:
            qs_frozen = qs_frozen.filter(scenario=active_scenario)
        else:
            qs_frozen = qs_frozen.filter(scenario__isnull=True)
            
        frozen_projects = list(qs_frozen.exclude(
            proyecto_code__isnull=True
        ).exclude(proyecto_code='').values_list(
            'proyecto_code', flat=True
        ).distinct())
        
        scenario_name = active_scenario.nombre if active_scenario else 'Sin Escenario'
        print(f"DEBUG: [Snapshot] Escenario '{scenario_name}' — {updated} tareas congeladas. Proyectos: {frozen_projects}")
        
        return JsonResponse({
            'status': 'ok',
            'tasks_frozen': updated,
            'proyectos': frozen_projects,
            'scenario_id': active_scenario.id if active_scenario else None,
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def hide_task(request):

    """
    API to hide a task from the list (virtual delete).
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        id_orden = body.get('id_orden')
        active_scenario = get_active_scenario(request)
        
        if not id_orden:
             return JsonResponse({'error': 'Missing ID'}, status=400)
             
        HiddenTask.objects.using('default').update_or_create(
            id_orden=id_orden, 
            scenario=active_scenario
        )
        
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def reactivar_op(request):
    """
    API to restore a hidden task (remove from hidden_task table).
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        id_orden = body.get('id_orden')
        # We can extract scenario_id directly to avoid redundant calls
        scenario_id = body.get('scenario_id')
        active_scenario = get_active_scenario(request, scenario_id=scenario_id)
        
        if not id_orden:
             return JsonResponse({'error': 'Missing ID'}, status=400)
             
        # Normalize ID
        try:
            id_orden_clean = int(float(id_orden))
        except:
            id_orden_clean = id_orden

        deleted_count, _ = HiddenTask.objects.using('default').filter(
            id_orden=id_orden_clean, 
            scenario=active_scenario
        ).delete()
        
        return JsonResponse({'status': 'ok', 'active': True, 'deleted_count': deleted_count})
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def update_manual_time(request):
    """
    API to update the manual process time for a task.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        id_orden = body.get('id_orden')
        tiempo_manual = body.get('tiempo_manual')
        maquina = body.get('maquina') or 'SIN ASIGNAR'
        active_scenario = get_active_scenario(request)
        
        if not id_orden or tiempo_manual is None:
             return JsonResponse({'error': 'Missing parameters'}, status=400)
             
        time_val = float(tiempo_manual)
        
        # We need to maintain the same machine for the override to be found correctly later
        obj, created = PrioridadManual.objects.using('default').get_or_create(
            id_orden=id_orden,
            scenario=active_scenario,
            maquina=maquina
        )
        obj.tiempo_manual = time_val
        obj.save(using='default')
            
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        print(f"ERROR update_manual_time: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def update_cantidad_producida(request):
    """
    API to update the produced quantity for a task manually.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    
    try:
        body = json.loads(request.body)
        id_orden = body.get("id_orden")
        cantidad_producida = body.get("cantidad_producida")
        maquina = body.get("maquina")
        scenario_id = body.get("scenario_id")
        
        if not id_orden or cantidad_producida is None:
            return JsonResponse({"error": "Missing parameters"}, status=400)
        
        active_scenario = get_active_scenario(request)
        if scenario_id:
            try:
                from .models import Scenario
                active_scenario = Scenario.objects.using('default').get(id=scenario_id)
            except: pass
            
        from django.db import transaction
        with transaction.atomic(using="default"):
            # Ensure we update the right record
            p = PrioridadManual.objects.using('default').filter(
                id_orden=id_orden, 
                scenario=active_scenario
            ).first()
            
            if not p:
                if not maquina:
                     return JsonResponse({"error": "No manual state found for this OP. Move it or change machine first."}, status=400)
                
                p = PrioridadManual.objects.using('default').create(
                    id_orden=id_orden,
                    maquina=maquina,
                    prioridad=0, 
                    scenario=active_scenario
                )
            
            p.cantidad_producida_manual = float(cantidad_producida)
            p.save(using='default')
            
        return JsonResponse({"status": "ok", "new_value": p.cantidad_producida_manual})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def update_manual_nivel(request):
    """
    API to update the manual planning level (Nivel Planificacion) for a task.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        id_orden = body.get('id_orden')
        nivel_manual = body.get('nivel_manual')
        maquina = body.get('maquina') or 'SIN ASIGNAR'
        active_scenario = get_active_scenario(request)
        
        if not id_orden or nivel_manual is None:
             return JsonResponse({'error': 'Missing parameters'}, status=400)
             
        nivel_val = int(nivel_manual)
        
        obj, created = PrioridadManual.objects.using('default').get_or_create(
            id_orden=id_orden,
            scenario=active_scenario,
            maquina=maquina
        )
        obj.nivel_manual = nivel_val
        obj.save(using='default')
            
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        print(f"ERROR update_manual_nivel: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def update_overlap_percentage(request):
    """
    API to update the overlap percentage for a task.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        id_orden = body.get('id_orden')
        porcentaje_solapamiento = body.get('porcentaje_solapamiento')
        modo_solapamiento = body.get('modo_solapamiento')
        maquina = body.get('maquina') or 'SIN ASIGNAR'
        active_scenario = get_active_scenario(request)
        
        if not id_orden:
            return JsonResponse({'error': 'Missing parameters'}, status=400)
            
        obj, created = PrioridadManual.objects.using('default').get_or_create(
            id_orden=id_orden,
            scenario=active_scenario,
            maquina=maquina
        )
        
        if porcentaje_solapamiento is not None:
            porcentaje_val = float(porcentaje_solapamiento)
            if porcentaje_val < 0 or porcentaje_val > 100:
                return JsonResponse({'error': 'Percentage must be between 0 and 100'}, status=400)
            obj.porcentaje_solapamiento = porcentaje_val
            
        if modo_solapamiento is not None:
            if modo_solapamiento not in ['manual', 'automatico']:
                return JsonResponse({'error': 'Invalid modo_solapamiento'}, status=400)
            obj.modo_solapamiento = modo_solapamiento
            
        obj.save(using='default')
        
        return JsonResponse({
            'status': 'ok',
            'id_orden': id_orden,
            'porcentaje_solapamiento': obj.porcentaje_solapamiento,
            'modo_solapamiento': obj.modo_solapamiento
        })
    except Exception as e:
        print(f"❌ ERROR update_overlap_percentage: {e}")
        return JsonResponse({'error': str(e)}, status=500)



from .planning_service import calculate_timeline
from datetime import datetime, timedelta
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
import json


def main_menu(request):
    return render(request, 'produccion/menu.html')

def planificacion_list(request):
    """
    View to retrieve planning data and render it in a table.
    """
    # Canonical URL Redirect: Ensure scenario_id is always in URL for consistency
    if 'scenario_id' not in request.GET:
        active_scenario = get_active_scenario(request)
        if active_scenario:
            params = request.GET.copy()
            params['scenario_id'] = active_scenario.id
            return redirect(f"{request.path}?{params.urlencode()}")

    active_scenario = get_active_scenario(request)
    print(f"DEBUG VISTA PRINCIPAL: Extrayendo procesos para el Scenario ID: {active_scenario.id if active_scenario else 'NONE'}")

    # Filter Persistence: Session + GET
    id_orden = request.GET.get('id_orden')
    if id_orden is not None:
        request.session['last_id_orden_filter'] = id_orden
    elif 'id_orden' not in request.GET:
        id_orden = request.session.get('last_id_orden_filter')

    # Strict Filtering: Only use projects from the current GET request
    import urllib.parse
    proyectos = request.GET.get('proyectos', '')
    if proyectos:
        proyectos = urllib.parse.unquote(proyectos).replace('%2C', ',').strip()

    # Build filtros for the SQL query
    filtros = {}
    if id_orden:
        filtros['id_orden'] = id_orden
    if proyectos:
        filtros['proyectos'] = [p.strip() for p in proyectos.split(',') if p.strip()]
    if active_scenario:
        filtros['scenario_id'] = active_scenario.id


    try:
        # --- Local Machine Config Logic ---
        local_machines = MaquinaConfig.objects.using('default').all()
        using_local_config = local_machines.exists()
        
        if using_local_config:
            # Map: { ID: Name }
            machine_map = {m.id_maquina.strip(): m.nombre for m in local_machines}
            # List of objects for the template
            all_machines_list = [{'id': m.id_maquina.strip(), 'nombre': m.nombre} for m in local_machines]
            filtros['machine_ids'] = list(machine_map.keys())
        else:
            names = get_all_machines()
            # If no local config, Name IS the ID
            all_machines_list = [{'id': n, 'nombre': n} for n in names]
            machine_map = {n: n for n in names}

        # NEW: Selective planning. Only show tasks that are explicitly in PlannedTask for this scenario.
        # We fetch the IDs from SQLite but the data from SQL Server (ERP).
        planned_tasks_qs = PlannedTask.objects.using('default').filter(scenario=active_scenario).order_by('proyecto_code', 'id_orden')
        planned_ids = list(planned_tasks_qs.values_list('id_orden', flat=True))
        
        # MERGE LOGIC: Ensure that any project already in PlannedTask is included in the filter.
        # This prevents "A disappearing when searching for B".
        url_projs = [p.strip() for p in (proyectos or "").split(',') if p.strip()]
        planned_projs = list(planned_tasks_qs.values_list('proyecto_code', flat=True).distinct())
        
        # Combined set of projects (keeping URL order if possible)
        combined_projs = url_projs.copy()
        for p in planned_projs:
            if p and p not in combined_projs:
                combined_projs.append(p)
        

        if combined_projs:
            proyectos = ",".join(combined_projs)
            filtros['proyectos'] = combined_projs

        # We always set this filter to ensure only selected tasks appear
        filtros['id_orden_in'] = planned_ids

        # Optimization: Only fetch data if we have planned IDs OR we are looking for a specific ID
        # If the user is just "searching" for a project, we don't load ERP data into the table yet
        # (the frontend will open the Selector modal instead).
        search_active = bool(planned_ids or filtros.get('id_orden'))
        
        if search_active:
            # We set exclude_completed=False because if a task is explicitly in PlannedTask
            # (Selected via the new UI), we WANT to see it even if it's technically completed in ERP.
            data = get_planificacion_data(filtros, exclude_completed=False)
        else:
            data = []
        
        # --- CANONICAL MACHINE HARMONIZATION ---
        # Build maps to translate between Name and ID for overrides
        name_to_id = {m.nombre.strip().upper(): m.id_maquina.strip() for m in local_machines}
        id_to_name = {m.id_maquina.strip(): m.nombre.strip() for m in local_machines}
        
        # Determine plan mode
        current_plan_mode = request.GET.get('plan_mode') or request.POST.get('plan_mode') or 'manual'
        plan_mode = current_plan_mode
        request.session['last_plan_mode'] = current_plan_mode
        
        # --- FILTER HIDDEN TASKS FOR THIS SCENARIO ---
        # If we are in 'audit_mode', we include hidden tasks but mark them.
        audit_mode = request.GET.get('audit_mode') == '1'
        hidden_ids = set()
        if plan_mode != 'original':
            hidden_ids = set(HiddenTask.objects.using('default').filter(scenario=active_scenario).values_list('id_orden', flat=True))
            
            if audit_mode:
                # In audit mode, we keep them but flag them
                for item in data:
                    if item.get('Idorden') in hidden_ids:
                        item['is_hidden'] = True
            else:
                # Normal mode: filter them out
                if hidden_ids:
                    data = [d for d in data if d.get('Idorden') not in hidden_ids]

        # 0. Fetch PrioridadManual levels to merge into data
        planned_metadata = {}
        if active_scenario:
            pm_list = PrioridadManual.objects.using('default').filter(scenario=active_scenario).values('id_orden', 'nivel_manual')
            planned_metadata = {p['id_orden']: p['nivel_manual'] for p in pm_list if p['nivel_manual'] is not None}

        # 0b. Fetch ProyectoPrioridad to know the ordering of projects for this scenario
        proj_prio_map = {}  # { proyecto_code: prioridad_proyecto (int) }
        if active_scenario:
            pp_qs = ProyectoPrioridad.objects.using('default').filter(scenario=active_scenario).values('proyecto', 'prioridad')
            proj_prio_map = {str(pp['proyecto']).strip(): pp['prioridad'] for pp in pp_qs}

        # Fallback incremental prioridades para proyectos no configurados en BD
        # Así el orden jerárquico ascendente se mantiene incluso sin registros explícitos
        for idx, p in enumerate(combined_projs or []):
            p_clean = str(p).strip()
            if p_clean not in proj_prio_map:
                proj_prio_map[p_clean] = idx + 1




        # Fetch Local Priorities filtered by Scenario
        virtual_overrides = {}
        id_to_any_override = {}
        if active_scenario:
            prioridades_db = PrioridadManual.objects.using('default').filter(scenario=active_scenario).order_by('orden_secuencia')
            print(f"DEBUG: planificacion_list - Loading {prioridades_db.count()} overrides for Scenario {active_scenario.nombre}")
                           # Map for OVERRIDES: (id_orden, maquina_id) -> data
            # We harvest ALL manual attributes to ensure consistency
            virtual_overrides = {}
            id_to_any_override = {}
            
            for p in prioridades_db:
                oid = int(p.id_orden)
                mid = str(p.maquina).strip()
                node = {
                    'maquina': mid, 
                    'prioridad': p.prioridad,
                    'tiempo_manual': p.tiempo_manual,
                    'nivel_manual': p.nivel_manual,
                    'porcentaje_solapamiento': p.porcentaje_solapamiento,
                    'modo_solapamiento': p.modo_solapamiento or 'automatico',
                    'cantidad_producida_manual': p.cantidad_producida_manual,
                    'fecha_inicio_manual': p.fecha_inicio_manual,
                    'orden_secuencia': p.orden_secuencia
                }
                virtual_overrides[(oid, mid)] = node
                id_to_any_override[oid] = mid
            
            id_to_override = id_to_any_override
        else:
            print("[WARN] No Active Scenario found.")

        # 1. Calculate extra fields, assign Priority, and Normalize Machine Name
        # We start with a BASELINE PROIRITY based on the initial SQL sort order (Index).
        # This prevents "unmoved" items (Priority 0) from being jumped over by a moved item (Priority 1500).
        for idx, item in enumerate(data):
            # Normalizar ID de orden para que coincida con overrides
            try:
                t_id_val = int(float(item.get('Idorden')))
                item['Idorden'] = t_id_val # Actualizar en el item
            except:
                t_id_val = 0

            # Attach piece priority removed (unified with nivel_manual)

            # Update Machine Name based on Local Config if active
            native_code = str(item.get('Idmaquina', '')).strip()
            
            # A. Determine NATIVE Machine Name & ID
            if using_local_config:
                # HARMONIZATION: Translate ERP code to our Canonical ID
                # erp_code could be "MAC18" OR "VF2" depending on ERP version
                erp_code = str(item.get('Idmaquina', '')).strip()
                canonical_id = erp_code
                
                # Check if erp_code is actually a Name in our config
                if erp_code.upper() in name_to_id:
                    canonical_id = name_to_id[erp_code.upper()]
                
                current_machine_id = canonical_id
                current_machine_name = id_to_name.get(canonical_id, erp_code)
            else:
                current_machine_id = item.get('MAQUINAD', 'SIN ASIGNAR')
                current_machine_name = current_machine_id
            
            # B. Check for VIRTUAL OVERRIDE (Moved Task)
            t_id_val = int(item.get('Idorden'))
            override_node = None
            
            # Use Canonical ID for lookup
            m_lookup = str(current_machine_id).strip()
            
            keys_to_try = [(t_id_val, m_lookup)]
            
            for k in keys_to_try:
                if k in virtual_overrides:
                    override_node = virtual_overrides[k]
                    break
            
            if not override_node:
                # Cross-machine lookup (Moved items)
                if t_id_val in id_to_any_override:
                    target_m_id = id_to_any_override[t_id_val]
                    override_node = virtual_overrides.get((t_id_val, target_m_id))

            if override_node:
                target_machine_id = str(override_node['maquina']).strip()
                # CRÍTICO: La "Prioridad Pieza" (1, 2, 3...) se lee del campo 'nivel_manual'
                # del registro PrioridadManual. NO usar 'prioridad' (FloatField con default 0.0
                # usado únicamente para OrdenVisual). Esta convención la confirma services.py:280.
                pieza_priority_val = override_node.get('nivel_manual')

                current_machine_id = target_machine_id
                current_machine_name = id_to_name.get(target_machine_id, target_machine_id)

                item['OrdenVisual'] = float(override_node.get('prioridad', 0))
                item['OrdenSecuencia'] = float(override_node.get('orden_secuencia', 999999))
                item['ManualPriorityFlag'] = True

                # nivel_planificacion: SIEMPRE el valor nativo del ERP (columna NIVEL de TMAN002).
                # NUNCA sobrescribir con nivel_manual. Eso es prioridad_articulo.
                item['NivelManualFlag'] = bool(override_node.get('nivel_manual') is not None)

                # Asignación de la Prioridad de la pieza (1, 2, 3...) — desde nivel_manual, NO desde prioridad
                if pieza_priority_val is not None:
                    item['prioridad_pieza'] = int(float(pieza_priority_val))
                else:
                    item['prioridad_pieza'] = item.get('prioridad_pieza', 999)

                if override_node.get('tiempo_manual') is not None:
                    item['Tiempo_Proceso'] = float(override_node['tiempo_manual'])
                    item['CalculadoManual'] = True
                else:
                    item['CalculadoManual'] = False

                if override_node.get('porcentaje_solapamiento') is not None:
                    item['porcentaje_solapamiento'] = override_node['porcentaje_solapamiento']
                if override_node.get('modo_solapamiento') is not None:
                    item['modo_solapamiento'] = override_node['modo_solapamiento']
                else:
                    item['modo_solapamiento'] = 'automatico'
            else:
                item['OrdenVisual'] = None
                item['ManualPriorityFlag'] = False
                item['CalculadoManual'] = False
                item['NivelManualFlag'] = False
                item['modo_solapamiento'] = 'automatico'

            # Final Assignment to Item
            item['MAQUINAD'] = current_machine_name
            item['MAQUINA_ID'] = current_machine_id

            # Si no existía el nodo de override, garantizamos los valores por defecto del item
            if not override_node:
                item['prioridad_pieza'] = item.get('prioridad_pieza', 999)
                item['modo_solapamiento'] = 'automatico'

            # nivel_planificacion: SIEMPRE el nativo del ERP (TMAN002), nunca el override manual.
            erp_nivel = item.get('Nivel_Planificacion')
            item['nivel_planificacion'] = int(erp_nivel) if erp_nivel is not None else 0

            p_code_clean = str(item.get('ProyectoCode') or '').strip()
            # BUG FIX: Distinguimos "registro explícito en ProyectoPrioridad" de "proyecto
            # sin prioridad asignada". Si NO existe registro, dejamos None en vez de
            # inyectar 999 (default del modelo), para que la celda no muestre el resguardo.
            # Inyectamos el valor bajo la clave EXACTA 'prioridad_del_proyecto' que el
            # template espera en la celda (alineado con el patrón annotate/F del ORM).
            if p_code_clean and p_code_clean in proj_prio_map:
                p_prio = proj_prio_map[p_code_clean]
            else:
                p_prio = None
            item['prioridad_del_proyecto'] = p_prio
            item['prioridad_proyecto'] = p_prio
            item['proyecto_prioridad'] = p_prio
            item['prioridad_articulo'] = item.get('prioridad_pieza')

            # Forzar que 'prioridad' apunte a la prioridad de la pieza (para compatibilidad con el template HTML)
            item['prioridad'] = item['prioridad_pieza']

            # Cantidades
            item['Cantidad'] = item.get('cantidad_final') or 0
            
            # Apply Manual overrides for quantity if exists
            manual_qty = None
            if override_node and override_node.get('cantidad_producida_manual') is not None:
                manual_qty = override_node['cantidad_producida_manual']
                item['Cantidadpp'] = manual_qty
                item['CantidadManualFlag'] = True
            else:
                item['Cantidadpp'] = item.get('cantidad_producida') or 0
                item['CantidadManualFlag'] = False

            item['CantidadesPendientes'] = max(0, item['Cantidad'] - item['Cantidadpp'])

            # Auditoria de Tiempos y Desvios (KPI)
            t_std = float(item.get('Tiempo') or 0.0)
            t_fichado_total = float(item.get('Total_Horas_Fichadas') or 0.0)
            c_prod = float(item['Cantidadpp'])
            
            t_real_unitario = 0.0
            if c_prod > 0:
                t_real_unitario = t_fichado_total / c_prod
                
            desvio_pct = 0.0
            if t_std > 0 and t_real_unitario > 0:
                desvio_pct = ((t_real_unitario - t_std) / t_std) * 100.0
                
            item['Tiempo_Real_Unitario'] = t_real_unitario
            item['Desvio_Porcentaje'] = desvio_pct
            
            if t_real_unitario <= 0 or t_std <= 0:
                item['KPI_Eficiencia'] = 'gray'
            elif desvio_pct <= 0:
                item['KPI_Eficiencia'] = 'green'
            elif desvio_pct <= 15.0:
                item['KPI_Eficiencia'] = 'yellow'
            elif desvio_pct <= 20.0:
                item['KPI_Eficiencia'] = 'orange'
            else:
                item['KPI_Eficiencia'] = 'red'


        # 2. Initialize Grouping using MACHINE NAMES
        grouped_data = {m['nombre']: [] for m in all_machines_list}
        if 'SIN ASIGNAR' not in grouped_data:
             grouped_data['SIN ASIGNAR'] = []        # Populate with data
        for item in data:
            m_name = item.get('MAQUINAD', 'SIN ASIGNAR')
            if m_name in grouped_data:
                grouped_data[m_name].append(item)
            else:
                if using_local_config:
                     if 'SIN ASIGNAR' in grouped_data:
                          grouped_data['SIN ASIGNAR'].append(item)
                else:
                    grouped_data.setdefault(m_name, []).append(item)
                    if m_name not in all_machines_list:
                        all_machines_list.append(m_name)
        
        # Sort items within each machine and re-assign visual IDs
        # Sort items within each machine and re-assign visual IDs
        for m_name in grouped_data:
            machine_items = grouped_data[m_name]
            
            # 1. Fill defaults for items without manual priority
            for idx, m_item in enumerate(machine_items):
                if m_item['OrdenVisual'] is None:
                    m_item['OrdenVisual'] = (idx + 1) * 5000.0 # Wide spacing for default

            def obtener_clave_ordenamiento_local(op):
                # Prioridad Proyecto
                p_proy = op.get('prioridad_del_proyecto') if isinstance(op, dict) else getattr(op, 'prioridad_del_proyecto', None)
                if p_proy is None or str(p_proy).strip() in ['—', '', 'None']:
                    p_proy_val = 999999
                else:
                    try:
                        p_proy_val = int(str(p_proy).strip())
                    except (ValueError, TypeError):
                        p_proy_val = 999999

                # Prioridad Artículos (Piezas)
                p_art = op.get('prioridad_articulo') if isinstance(op, dict) else getattr(op, 'prioridad_articulo', None)
                if p_art is None or str(p_art).strip() in ['—', '', 'None']:
                    p_art_val = 999999
                else:
                    try:
                        p_art_val = int(str(p_art).strip())
                    except (ValueError, TypeError):
                        p_art_val = 999999

                # Nivel Planificación (Procesos)
                n_plan = op.get('nivel_planificacion') if isinstance(op, dict) else getattr(op, 'nivel_planificacion', None)
                if n_plan is None or str(n_plan).strip() in ['—', '', 'None']:
                    n_plan_val = 999999
                else:
                    try:
                        n_plan_val = int(str(n_plan).strip())
                    except (ValueError, TypeError):
                        n_plan_val = 999999

                return (p_proy_val, p_art_val, -n_plan_val, op.get('OrdenSecuencia', 999999), op.get('Idorden', 9999999))

            # 2. Jerarquía estricta: Prioridad Proyecto (ASC) → Prioridad Pieza (ASC) → Nivel Planif. (DESC) → OrdenSecuencia (ASC) → IdOrden (ASC)
            # Solo si el usuario NO tiene un orden visual previo guardado en BD, aplicamos el local.
            # Ordenamos la lista localmente solo para dar valores por defecto a los que no tienen OrdenVisual.
            machine_items_no_manual = [m for m in machine_items if not m.get('ManualPriorityFlag', False)]
            machine_items_no_manual.sort(key=obtener_clave_ordenamiento_local)
            
            for idx, m_item in enumerate(machine_items_no_manual):
                # Les damos un OrdenVisual alto para que queden al final del bloque manual
                m_item['OrdenVisual'] = (idx + 1) * 50000.0

        # FINAL FILTER: REMOVED per user request ("no las ocultes")
        # We keep all machines visible to allow moving tasks to them.
        processed_machines = []
        for m in all_machines_list:
             processed_machines.append({'id': m['id'], 'nombre': m['nombre']})
        
        # Sort by name
        processed_machines.sort(key=lambda x: x['nombre'])

        if any(m['nombre'] == 'SIN ASIGNAR' for m in processed_machines):
             # Ensure SIN ASIGNAR is at the end
             sin_a = [m for m in processed_machines if m['nombre'] == 'SIN ASIGNAR'][0]
             processed_machines.remove(sin_a)
             processed_machines.append(sin_a)
        elif 'SIN ASIGNAR' in grouped_data:
             processed_machines.append({'id': 'SIN ASIGNAR', 'nombre': 'SIN ASIGNAR'})
        
        # FINAL FILTER: REMOVED per user request ("no las ocultes")
        # We keep all machines visible to allow moving tasks to them.
        # if search_active:
        #      processed_machines = [m for m in processed_machines if grouped_data.get(m)]

        # =========================================================
        # INYECCIÓN AGRESIVA: cruce EXPLICITO entre la grilla y la tabla
        # de prioridades. La grilla expone el código como 'item.ProyectoCode'
        # (mismo string que muestra la columna "Proyecto" del HTML). La tabla
        # 'ProyectoPrioridad' guarda el código en el campo CharField 'proyecto'.
        # Equivalencia: item.ProyectoCode == ProyectoPrioridad.proyecto
        # No confiar en inyecciones tempranas (se pierden en agrupación).
        # Usa el modelo importado al inicio del módulo (línea 10) — NUNCA inline
        # para no generar UnboundLocalError.
        # =========================================================
        pp_qs_last = ProyectoPrioridad.objects.using('default').filter(
            scenario=active_scenario
        ).values('proyecto', 'prioridad')
        full_prio_map = {str(p['proyecto']).strip(): p['prioridad'] for p in pp_qs_last}

        for m_name, machine_items in grouped_data.items():
            for item in machine_items:
                try:
                    # 1) Resolver el código del proyecto del item (dict u objeto)
                    if isinstance(item, dict):
                        cod_proyecto = item.get('ProyectoCode')
                    else:
                        cod_proyecto = getattr(item, 'ProyectoCode', None)

                    # 2) Cruzar contra la tabla de prioridades
                    val_prio = "—"
                    if cod_proyecto:
                        key = str(cod_proyecto).strip()
                        # Primero: mapa pre-construido (evita N+1)
                        if key in full_prio_map:
                            val_prio = str(full_prio_map[key])
                        else:
                            # Fallback: query fresca por si el mapa se construyó
                            # con otro escenario o el campo no se cacheó bien.
                            # El campo del FK es 'scenario', NO 'active_scenario'.
                            prio_registro = ProyectoPrioridad.objects.filter(
                                proyecto=key,
                                scenario=active_scenario
                            ).first()
                            if prio_registro:
                                val_prio = str(prio_registro.prioridad)
                except Exception:
                    val_prio = "—"

                # 3) Setear la variable EXACTA que el template lee
                # Redundante en dict y objeto para anular fallos de tipado
                if isinstance(item, dict):
                    item['prioridad_del_proyecto'] = val_prio
                    item['prioridad_proyecto'] = val_prio
                    item['proyecto_prioridad'] = val_prio
                else:
                    setattr(item, 'prioridad_del_proyecto', val_prio)
                    setattr(item, 'prioridad_proyecto', val_prio)
                    setattr(item, 'proyecto_prioridad', val_prio)

        def obtener_clave_ordenamiento(op, mode='manual'):
            id_orden = op.get('Idorden')
            manual_mode_activo = str(mode).strip().lower() == 'manual'

            if manual_mode_activo:
                orden_visual = op.get('OrdenVisual', 1000.0)
                try:
                    orden_visual_val = float(orden_visual)
                except (ValueError, TypeError):
                    orden_visual_val = 1000.0
                
                # Limitamos los logs a los primeros 5 por cuestión de spam
                if getattr(obtener_clave_ordenamiento, 'debug_count', 0) < 15:
                    print(f"[DEBUG LECTURA] OP {id_orden} | Mode: {mode} | OrdenVisual leído: {orden_visual}")
                    obtener_clave_ordenamiento.debug_count = getattr(obtener_clave_ordenamiento, 'debug_count', 0) + 1
                    
                return (0, orden_visual_val, id_orden)

            # MODO AUTOMÁTICO (Queda exactamente igual que antes)
            p_proy = op.get('prioridad_del_proyecto') if isinstance(op, dict) else getattr(op, 'prioridad_del_proyecto', None)
            if p_proy is None or str(p_proy).strip() in ['—', '', 'None']:
                p_proy_val = 999999
            else:
                try:
                    p_proy_val = int(str(p_proy).strip())
                except (ValueError, TypeError):
                    p_proy_val = 999999

            p_art = op.get('prioridad_articulo') if isinstance(op, dict) else getattr(op, 'prioridad_articulo', None)
            if p_art is None or str(p_art).strip() in ['—', '', 'None']:
                p_art_val = 999999
            else:
                try:
                    p_art_val = int(str(p_art).strip())
                except (ValueError, TypeError):
                    p_art_val = 999999

            n_plan = op.get('nivel_planificacion') if isinstance(op, dict) else getattr(op, 'nivel_planificacion', None)
            if n_plan is None or str(n_plan).strip() in ['—', '', 'None']:
                n_plan_val = 999999
            else:
                try:
                    n_plan_val = int(str(n_plan).strip())
                except (ValueError, TypeError):
                    n_plan_val = 999999

            return (2, p_proy_val, p_art_val, -n_plan_val, op.get('OrdenSecuencia', 999999), id_orden)

        # Forzar ordenamiento en cualquier variante estructural de grouped_data
        if isinstance(grouped_data, dict):
            for maquina, lista_ops in grouped_data.items():
                print(f"DEBUG ORDEN: Ordenando lista de ops para máquina {maquina} ({len(lista_ops)} items) | mode={current_plan_mode}")
                print(f"DEBUG ORDEN: Antes de ordenar, IDs: {[op.get('Idorden') for op in lista_ops]}")
                lista_ops.sort(key=lambda op: obtener_clave_ordenamiento(op, current_plan_mode))
                print(f"DEBUG ORDEN: Después de ordenar, IDs: {[op.get('Idorden') for op in lista_ops]}")
        elif isinstance(grouped_data, list):
            for item_maquina in grouped_data:
                if isinstance(item_maquina, dict):
                    if 'operaciones' in item_maquina and isinstance(item_maquina['operaciones'], list):
                        item_maquina['operaciones'].sort(key=lambda op: obtener_clave_ordenamiento(op, current_plan_mode))
                    if 'items' in item_maquina and isinstance(item_maquina['items'], list):
                        item_maquina['items'].sort(key=lambda op: obtener_clave_ordenamiento(op, current_plan_mode))
            # Por si es una lista plana de operaciones:
            try:
                grouped_data.sort(key=lambda op: obtener_clave_ordenamiento(op, current_plan_mode))
            except Exception:
                pass

        # Determine response format
        if request.GET.get('format') == 'json':
             return JsonResponse({'data': data}, safe=False)

        if isinstance(grouped_data, dict):
            print("DEBUG CONTEXTO TEMPLATE: orden final por máquina enviado a planificacion.html")
            for maquina, lista_ops in grouped_data.items():
                print(f"  Máquina {maquina}: {[op.get('Idorden') for op in lista_ops]}")

        return render(request, 'produccion/planificacion.html', {
            'grouped_data': grouped_data, 
            'machines': processed_machines,
            'search_active': search_active,
            'proyectos_value': proyectos if proyectos else '',
            'id_orden_value': id_orden if id_orden else '',
            'all_scenarios': Scenario.objects.using('default').all() if 'Scenario' in globals() else [],
            'active_scenario': active_scenario,
            'active_scenario_id': active_scenario.id if active_scenario else None,
            'audit_mode': audit_mode
        })
    except Exception as e:
        if request.GET.get('format') == 'json':
            return JsonResponse({'error': str(e)}, status=500)
        return render(request, 'produccion/planificacion.html', {'grouped_data': {}, 'machines': [], 'error': str(e)})


@csrf_exempt
@csrf_exempt
def move_priority(request, id_orden, direction):
    """
    API to move an order up or down in the local priority list.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        maquina_raw = body.get('maquina')
        # SANITIZE: replace comma decimal separator before float conversion
        current_priority = float(str(body.get('priority', 0)).replace(',', '.'))
        neighbor_id = body.get('neighbor_id')
        neighbor_priority = float(str(body.get('neighbor_priority', 0)).replace(',', '.'))
        active_scenario = get_active_scenario(request)
        
        if neighbor_id is None:
            return JsonResponse({'status': 'ignored', 'message': 'No neighbor'})

        # Harmonize machine key
        from .models import MaquinaConfig
        m_conf = MaquinaConfig.objects.filter(nombre=maquina_raw).first()
        maquina_id = m_conf.id_maquina if m_conf else maquina_raw

        # Target Item (Delete old name-based if moving to ID-based)
        if m_conf and maquina_raw != maquina_id:
             PrioridadManual.objects.filter(id_orden=id_orden, maquina=maquina_raw, scenario=active_scenario).delete()
             PrioridadManual.objects.filter(id_orden=neighbor_id, maquina=maquina_raw, scenario=active_scenario).delete()

        obj_target, _ = PrioridadManual.objects.using("default").update_or_create(
            id_orden=id_orden, maquina=maquina_id, scenario=active_scenario,
            defaults={"prioridad": neighbor_priority}
        )
        
        obj_neighbor, _ = PrioridadManual.objects.using("default").update_or_create(
            id_orden=neighbor_id, maquina=maquina_id, scenario=active_scenario,
            defaults={"prioridad": current_priority}
        )
        
        return JsonResponse({"status": "ok"})
    except Exception as e:
        return JsonResponse({"error": f"DB Error: {str(e)}"}, status=500)

@csrf_exempt
def move_task(request):
    """
    API to move a task to a different machine and/or update its priority order.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    
    try:
        body = json.loads(request.body)
        id_orden = body.get("id_orden")
        target_machine_raw = body.get("target_machine_id")
        new_priority = body.get("new_priority")
        active_scenario = get_active_scenario(request)
        
        print(f"--- DETECCIÓN: Intentando mover OP {id_orden} (type={type(id_orden)}) a la Máquina {target_machine_raw} (type={type(target_machine_raw)}), Prio={new_priority} (type={type(new_priority)}) ---")
        
        if id_orden is None or target_machine_raw is None or str(target_machine_raw).strip() == '' or new_priority is None:
             print(f"--- MOVE_TASK RECHAZADO: id_orden={id_orden}, target_machine_raw={target_machine_raw}, new_priority={new_priority} ---")
             return JsonResponse({"error": f"Missing parameters: id_orden={id_orden}, target_machine_id={target_machine_raw}, new_priority={new_priority}"}, status=400)
             
        new_priority = float(str(new_priority).replace(',', '.'))
        
        # Harmonize machine: look up by id_maquina first (what the tab sends now),
        # then by nombre, then fall back to raw value as-is.
        from .models import MaquinaConfig
        print(f"DEBUG: move_task - ID: {id_orden}, TargetRaw: {target_machine_raw}, Prio: {new_priority}")
        m_conf = MaquinaConfig.objects.using("default").filter(id_maquina=target_machine_raw).first()
        if not m_conf:
            m_conf = MaquinaConfig.objects.using("default").filter(nombre=target_machine_raw).first()
        target_machine_id = m_conf.id_maquina if m_conf else target_machine_raw
        print(f"DEBUG: move_task - Resolved Machine ID: {target_machine_id}")


        from django.db import transaction
        with transaction.atomic(using="default"):
            # Normalizar ID (SQL vs Django)
            try:
                id_orden_clean = int(float(id_orden))
            except:
                id_orden_clean = id_orden

            # Fetch existing to preserve attributes
            old_entry = PrioridadManual.objects.using("default").filter(id_orden=id_orden_clean, scenario=active_scenario).first()
            
            existing_data = {
                "tiempo_manual": old_entry.tiempo_manual if old_entry else None,
                "fecha_inicio_manual": old_entry.fecha_inicio_manual if old_entry else None,
                "nivel_manual": old_entry.nivel_manual if old_entry else None,
                "porcentaje_solapamiento": old_entry.porcentaje_solapamiento if old_entry else 0.0,
                "modo_solapamiento": "manual"  # Force to manual upon drag/drop re-assignment
            }
            
            # Clean up all assignments for this OP in this scenario
            PrioridadManual.objects.using("default").filter(id_orden=id_orden_clean, scenario=active_scenario).delete()
            
            op = PrioridadManual.objects.using("default").create(
                id_orden=id_orden_clean,
                maquina=target_machine_id,
                prioridad=new_priority,
                scenario=active_scenario,
                **existing_data
            )
            print(f"--- BASE DE DATOS: OP {id_orden_clean} guardada exitosamente en máquina {op.maquina} ---")
            
            current_plan_mode = request.GET.get('plan_mode') or request.POST.get('plan_mode')
            if not current_plan_mode:
                current_plan_mode = request.session.get('last_plan_mode', 'manual')
            request.session['last_plan_mode'] = current_plan_mode
            
        return JsonResponse({"status": "ok"})
    except Exception as e:
        return JsonResponse({"error": f"DB Error: {str(e)}"}, status=500)

@csrf_exempt
def set_priority(request, id_orden):
    """
    API to set a specific priority AND/OR manual start date for an order.
    Used for Drag and Drop (pinning/manual sequencing).
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        maquina = body.get('maquina')
        new_priority = body.get('new_priority')
        manual_start_str = body.get('manual_start')
        active_scenario = get_active_scenario(request)
        
        # Validation
        if new_priority is None and manual_start_str is None:
             return JsonResponse({'error': 'Missing new_priority or manual_start'}, status=400)
             
        if new_priority is not None:
            new_priority = float(str(new_priority).replace(',', '.'))
            
        manual_start_dt = None
        if manual_start_str:
            try:
                from django.utils import timezone as django_tz
                # Robust Date Parsing
                manual_start_str = str(manual_start_str).strip()
                if 'T' in manual_start_str:
                    manual_start_dt = datetime.fromisoformat(manual_start_str.replace('Z', '+00:00'))
                    if manual_start_dt.tzinfo is None:
                        manual_start_dt = django_tz.make_aware(manual_start_dt)
                else:
                    if '.' in manual_start_str:
                        manual_start_str = manual_start_str.split('.')[0]
                    naive_dt = datetime.strptime(manual_start_str, '%Y-%m-%d %H:%M:%S')
                    manual_start_dt = django_tz.make_aware(naive_dt)
            except Exception as ve:
                print(f"ERROR parsing date in set_priority: {ve}")
                return JsonResponse({'error': f'Invalid date format: {manual_start_str}'}, status=400)

        # Harmonize machine
        from .models import MaquinaConfig
        print(f"DEBUG: set_priority - ID: {id_orden}, TargetRaw: {maquina}, Prio: {new_priority}, Scenario: {active_scenario.id if active_scenario else 'None'}")
        m_conf = MaquinaConfig.objects.using('default').filter(nombre=maquina).first()
        maquina_id = m_conf.id_maquina if m_conf else maquina
        print(f"DEBUG: set_priority - Resolved Machine ID: {maquina_id}")


        from django.db import transaction
        with transaction.atomic(using='default'):
            # Normalizar ID para asegurar match con DB (SQL vs Django types)
            try:
                id_orden_clean = int(float(id_orden))
            except:
                id_orden_clean = id_orden
                
            # Fetch existing to preserve ALL other manual overrides
            old_entry = PrioridadManual.objects.using('default').filter(id_orden=id_orden_clean, scenario=active_scenario).first()
            
            # Default values if no entry exists
            existing_data = {
                'tiempo_manual': old_entry.tiempo_manual if old_entry else None,
                'nivel_manual': old_entry.nivel_manual if old_entry else None,
                'porcentaje_solapamiento': old_entry.porcentaje_solapamiento if old_entry else 0.0,
                'fecha_inicio_manual': old_entry.fecha_inicio_manual if old_entry else None,
                'prioridad': old_entry.prioridad if old_entry else (new_priority if new_priority is not None else 0.0),
                'modo_solapamiento': 'manual'  # Force to manual on Gantt drag/drop/reorder
            }
            
            # Clean up before re-creating
            PrioridadManual.objects.using('default').filter(id_orden=id_orden_clean, scenario=active_scenario).delete()
            
            final_start_date = manual_start_dt if manual_start_dt is not None else existing_data['fecha_inicio_manual']
            final_priority = new_priority if new_priority is not None else existing_data['prioridad']
            
            PrioridadManual.objects.using('default').create(
                id_orden=id_orden_clean,
                maquina=maquina_id, 
                prioridad=final_priority,
                fecha_inicio_manual=final_start_date,
                scenario=active_scenario,
                tiempo_manual=existing_data['tiempo_manual'],
                nivel_manual=existing_data['nivel_manual'],
                porcentaje_solapamiento=existing_data['porcentaje_solapamiento'],
                modo_solapamiento=existing_data['modo_solapamiento']
            )

            current_plan_mode = request.GET.get('plan_mode') or request.POST.get('plan_mode')
            if not current_plan_mode:
                current_plan_mode = request.session.get('last_plan_mode', 'manual')
            request.session['last_plan_mode'] = current_plan_mode
            
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        print(f"❌ ERROR set_priority DB: {e}")
        return JsonResponse({'error': f'DB Error: {str(e)}'}, status=500)
    




# --- Machine Configuration Views ---

from django.db import transaction

from django.core.paginator import Paginator

def maquina_config_list(request):
    """
    List all locally configured machines and their schedules + Equivalencies.
    """
    from .models import MaquinaEquivalencia
    
    # Order by ID to ensure consistent pagination
    maquinas_list = MaquinaConfig.objects.using('default').prefetch_related('horarios').all().order_by('id_maquina')
    all_maquinas = MaquinaConfig.objects.using('default').all().order_by('id_maquina')
    
    paginator = Paginator(maquinas_list, 10) # Increased to 10
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Equivalencies for the management section
    equivalencias = MaquinaEquivalencia.objects.using('default').select_related('maquina_origen', 'maquina_destino').all()
    
    return render(request, 'produccion/maquina_config_list.html', {
        'maquinas': page_obj,
        'all_maquinas': all_maquinas,
        'equivalencias': equivalencias
    })

@csrf_exempt
def maquina_equivalencia_save(request):
    """
    Save or delete a machine equivalency.
    """
    from .models import MaquinaEquivalencia, MaquinaConfig
    from django.shortcuts import redirect
    from django.contrib import messages
    
    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        
        if action == 'delete':
            eq_id = request.POST.get('id')
            MaquinaEquivalencia.objects.using('default').filter(pk=eq_id).delete()
            messages.success(request, "Equivalencia eliminada correctamente.")
        else:
            origen_id = request.POST.get('origen')
            destino_id = request.POST.get('destino')
            eficiencia = float(request.POST.get('eficiencia', 1.0))
            
            if origen_id == destino_id:
                messages.error(request, "La máquina origen y destino no pueden ser la misma.")
            else:
                origen = MaquinaConfig.objects.using('default').get(id_maquina=origen_id)
                destino = MaquinaConfig.objects.using('default').get(id_maquina=destino_id)
                
                MaquinaEquivalencia.objects.using('default').update_or_create(
                    maquina_origen=origen,
                    maquina_destino=destino,
                    defaults={'factor_eficiencia': eficiencia}
                )
                messages.success(request, f"Equivalencia {origen_id} -> {destino_id} guardada.")
                
    return redirect('maquina_config_list')

def planificacion_visual_OLD(request):
    """
    Visual Gantt Chart View.
    """
    # 1. Get Local Machines
    maquinas = MaquinaConfig.objects.using('default').prefetch_related('horarios').all().order_by('id_maquina')
    
    # 2. Get Data and Calculate Timeline
    timeline_data = [] 
    
    # 3. Check for Local Manual Priorities/Group Assignments & Time Overrides
    active_scenario = get_active_scenario(request)

    # 2. Prepare Start Date for Simulation
    # Parse fecha_desde if provided, otherwise use scenario's start date, otherwise use today
    fecha_desde = request.GET.get('fecha_desde')
    if fecha_desde:
        try:
            start_simulation = datetime.strptime(fecha_desde, '%Y-%m-%d')
        except ValueError:
            start_simulation = datetime.now()
    elif active_scenario and active_scenario.fecha_inicio:
        start_simulation = datetime.combine(active_scenario.fecha_inicio, datetime.min.time())
    else:
        start_simulation = datetime.now()
    
    # IMPORTANT: Start from beginning of workday (7:00 AM), not current time.
    # EXCEPTION: If planning TODAY, use MAX(hora_actual, 07:00) so tasks don't
    # get scheduled in the past. For future dates, always start at 07:00.
    workday_start = start_simulation.replace(hour=7, minute=0, second=0, microsecond=0)
    from datetime import date as _date
    if start_simulation.date() == _date.today():
        # Hoy: tomar la hora actual redondeada al próximo cuarto de hora, mínimo 07:00
        now_local = datetime.now()
        now_rounded = now_local.replace(second=0, microsecond=0)
        start_simulation = max(workday_start, now_rounded)
    else:
        # Día futuro: siempre arrancar a las 07:00
        start_simulation = workday_start

    # --------------------------------------------------------------------------
    # FILTER: Determine active projects EXCLUSIVELY from URL parameters.
    #
    # REGLA ABSOLUTA: El Gantt NO lee sesiones antiguas.
    # Si el usuario no pasó ?proyectos=... en la URL → pantalla vacía.
    # La sesión es propiedad del Tablero Azul, no del Gantt.
    # --------------------------------------------------------------------------
    proyectos_val = request.GET.get('proyectos', '').strip()
    proyectos_activos = [p.strip() for p in proyectos_val.split(',') if p.strip()]

    # HARD RESET: If table is empty or projects param is explicitly empty string, clear all cached data
    # This prevents the system from "remembering" old projects
    clear_flag = request.GET.get('clear', '0') == '1'
    if clear_flag or proyectos_val == '':
        # Signal frontend to clear visual cache
        request.session['gantt_needs_clear'] = True

    # Using PrioridadManual table to override positions and machines locally (Virtual Moves)
    # CRITICAL: Only pull overrides for the ACTIVE SCENARIO.
    virtual_overrides = {}
    manual_entries = PrioridadManual.objects.using('default').filter(scenario=active_scenario)
    for entry in manual_entries:
        virtual_overrides[entry.id_orden] = {
            'maquina': entry.maquina,
            'prioridad': entry.prioridad,
            'tiempo_manual': entry.tiempo_manual,
            'nivel_manual': entry.nivel_manual,
            'manual_start': entry.fecha_inicio_manual
        }

    # Create a set of IDs that are moved TO a machine locally
    tasks_moved_in_map = {}
    for oid, override_data in virtual_overrides.items():
        mid = override_data['maquina']
        if mid not in tasks_moved_in_map:
             tasks_moved_in_map[mid] = []
        tasks_moved_in_map[mid].append(oid)
    
    # --- HIDDEN TASKS ---
    # Fetch list of hidden task IDs for THIS SCENARIO ONLY to exclude them from the Gantt
    hidden_ids = set(HiddenTask.objects.using('default').filter(scenario=active_scenario).values_list('id_orden', flat=True))

    # EXECUTION MODE CHECK
    # User requested: "Cuando presione VER GANTT, solo ingrese... pero si procesar nada."
    # "EJECUTAR Gantt" button adds &run=1
    run_calculation = request.GET.get('run') == '1'

    if not run_calculation:
        # Return empty structure with machines but no tasks
        for maquina in maquinas:
            timeline_data.append({
                'machine': maquina,
                'tasks': []
            })
        
        # Skip all the complex logic below
        context = {
            'timeline_data': timeline_data,
            'today': start_simulation,
            'time_columns': range(7, 22), # Default columns for grid
            'total_width': 15 * 40,
            'dependencies_json': '[]',
        }
        return render(request, 'produccion/planificacion_visual.html', context)

    # --- AUTOMATIC DEPENDENCY PREPARATION (OPTION B: By Nivel Decreasing) ---
#    print("\n" + "=" * 70)
#    print("🔢 OPCIÓN B: Dependencias Automáticas por Nivel (Mayor a Menor)")
#    print("=" * 70)

    # -----------------------------------------------------------------------
    # PLANNED IDs - SQLite-native project filter via proyecto_code
    #
    # RULE: El Gantt es REACTIVO. Solo muestra lo que se buscó explícitamente.
    #       Si no hay proyectos activos → retorna vacio inmediatamente.
    #       NUNCA carga el historial acumulado por default.
    # -----------------------------------------------------------------------
    if not proyectos_activos:
        # Sin filtro de proyectos = pantalla vacía. Retornamos de inmediato.
        for maquina in maquinas:
            timeline_data.append({'machine': maquina, 'tasks': []})
        context = {
            'timeline_data': timeline_data,
            'today': start_simulation,
            'time_columns': range(7, 22),
            'total_width': 15 * 40,
            'dependencies_json': '[]',
            'plan_mode': (
                request.GET.get('plan_mode')
                or request.POST.get('plan_mode')
                or request.session.get('last_plan_mode', 'manual')
            ),
            'active_scenario_id': active_scenario.id if active_scenario else None,
            'all_scenarios': Scenario.objects.using('default').all(),
            'proyectos_value': proyectos_val or '',
            'gantt_empty_reason': 'no_projects',
            'gantt_needs_clear': True,
        }
        return render(request, 'produccion/planificacion_visual.html', context)

    # Hay proyectos activos: filtrar planned_ids estrictamente por proyecto_code en SQLite.
    planned_ids = list(
        PlannedTask.objects.using('default')
        .filter(scenario=active_scenario, proyecto_code__in=proyectos_activos)
        .values_list('id_orden', flat=True)
    )

    # Fallback para registros legacy sin proyecto_code: filtramos por ERP proyectos
    # pero NO cargamos todo el acumulado sin filtro.
    deps_filter = {}
    deps_filter['proyectos'] = proyectos_activos  # Siempre filtramos por proyecto activo

    if planned_ids:
        deps_filter['id_orden_in'] = planned_ids
    elif request.GET.get('id_orden'):
        deps_filter['id_orden'] = request.GET.get('id_orden')
        
    # search_active = True si hay proyectos activos (ya garantizado arriba)
    search_active = bool(planned_ids or proyectos_activos or request.GET.get('id_orden'))
    
    if search_active:
        all_tasks_for_deps = get_planificacion_data(deps_filter) 
    else:
        all_tasks_for_deps = []
    
    # DEBUG: Trace specific IDs
    debug_ids = [46762, 46759]
    for t in all_tasks_for_deps:
        if t.get('Idorden') in debug_ids:
             print(f"DEBUG TRACE {t.get('Idorden')}: Mstnmbr={t.get('Mstnmbr')}, Nivel={t.get('Nivel_Planificacion')}") 

    # Debug: Check if specific tasks are in virtual_overrides
    for task_id in [46543, 46542]:
        if task_id in virtual_overrides:
            print(f"DEBUG: Task {task_id} found in virtual_overrides: {virtual_overrides[task_id]}")
        else:
            print(f"DEBUG: Task {task_id} NOT found in virtual_overrides")
    
    # --- Apply Overrides to Dependency Candidates ---
    for task in all_tasks_for_deps:
        p_id = task.get('Idorden')
        ov_data = None
        # Robust Lookup
        if p_id in virtual_overrides:
             ov_data = virtual_overrides[p_id]
        else:
             try:
                 p_id_int = int(p_id)
                 if p_id_int in virtual_overrides:
                     ov_data = virtual_overrides[p_id_int]
             except: pass
        
        if ov_data and ov_data.get('nivel_manual') is not None:
             task['prioridad_pieza'] = ov_data['nivel_manual']
             # Debug: Show when manual nivel is applied
             if p_id in [46543, 46542]:
                 print(f"DEBUG OVERRIDE: Applied nivel_manual={ov_data['nivel_manual']} to task {p_id}")
             
    # Debug: Check Sample
    # print(f"DEBUG: Checking Mstnmbr/Nivel for deps. Count: {len(all_tasks_for_deps)}") 

    # 2. Group by Formula/ProyectoCode (NOT by Mstnmbr)
    # This ensures dependencies are created only between tasks of the same project
    from collections import defaultdict
    orders_map = defaultdict(list)
    
    for task in all_tasks_for_deps:
        formula = task.get('ProyectoCode')  # Changed from Mstnmbr to ProyectoCode
        # Only group if Formula exists
        if formula:
            orders_map[formula].append(task)

    dependency_map = {}
    dependencies_list_for_json = []

    for formula, tasks_in_order in orders_map.items():
        # Debug: Show all tasks and niveles for projects 25-100 and 25-098
        if formula in ['25-100', '25-098']:
            print(f"\n  DEBUG: Formula {formula} has {len(tasks_in_order)} tasks:")
            for t in tasks_in_order:
                nivel_p = t.get('Nivel_Planificacion')
                nivel = t.get('Nivel')
                print(f"    - Task {t.get('Idorden')}: Nivel_Planificacion={nivel_p}, Nivel={nivel} ({t.get('MAQUINAD')})")
        
        # Helper to get Nivel safely
        def get_nivel(t):
            try:
                # Use ONLY Nivel_Planificacion (not Nivel)
                val = t.get('Nivel_Planificacion')
                
                if val is None: 
                    return 0
                return float(val)
            except (ValueError, TypeError):
                return 0


        # Sort by Nivel DESCENDING (Highest Nivel = First Operation)
        tasks_sorted = sorted(tasks_in_order, key=get_nivel, reverse=True)
        
        # FILTER OUT tasks that are "SIN ASIGNAR" (not assigned to any machine)
        # We only want to create dependencies between tasks that are actually scheduled
        tasks_assigned = [t for t in tasks_sorted if t.get('MAQUINAD') and t.get('MAQUINAD') != 'SIN ASIGNAR']
        
        if formula in ['25-100', '25-098']:
            print(f"\n  DEBUG: After filtering SIN ASIGNAR, {len(tasks_assigned)} tasks remain:")
            for t in tasks_assigned:
                print(f"    - Task {t.get('Idorden')}: Nivel {get_nivel(t)} ({t.get('MAQUINAD')})")
        
        # Group tasks by nivel (only assigned tasks)
        nivel_groups = {}
        for task in tasks_assigned:
            nivel = get_nivel(task)
            if nivel not in nivel_groups:
                nivel_groups[nivel] = []
            nivel_groups[nivel].append(task)
        
        # Get sorted list of unique niveles (descending)
        sorted_niveles = sorted(nivel_groups.keys(), reverse=True)
        
        # Create dependencies: each nivel depends on the immediately higher nivel
        for i in range(len(sorted_niveles) - 1):
            higher_nivel = sorted_niveles[i]
            lower_nivel = sorted_niveles[i + 1]
            
            # All tasks in lower_nivel depend on ALL tasks in higher_nivel
            for successor in nivel_groups[lower_nivel]:
                succ_id = successor.get('Idorden')
                if not succ_id:
                    continue
                
                for predecessor in nivel_groups[higher_nivel]:
                    pred_id = predecessor.get('Idorden')
                    if not pred_id or pred_id == succ_id:
                        continue
                    
                    if succ_id not in dependency_map:
                        dependency_map[succ_id] = []
                    
                    # Avoid duplicates
                    if pred_id not in dependency_map[succ_id]:
                        dependency_map[succ_id].append(pred_id)
                        
                        dependencies_list_for_json.append({
                            'pred': pred_id,
                            'succ': succ_id
                        })
                        
                        # Debug logging for our specific tasks
                        if pred_id in [46762, 46759] or succ_id in [46762, 46759]:
                            print(f"  [DEPENDENCY CREATED] {pred_id} (Nivel {higher_nivel}) -> {succ_id} (Nivel {lower_nivel})")


#    print(f"\n  ✅ Created {len(dependency_map)} automatic dependencies based on Nivel (Desc)")
#    print("=" * 70 + "\n")
        
    # Global map to track end dates of ALL tasks across ALL machines
    global_task_end_dates = {}
    
    # Store machine data for second pass
    machine_tasks_map = {}  # machine_id -> {'maquina': obj, 'tasks': [...]}

    # ========================================================================
    # FIRST PASS: Calculate ALL tasks to build global_task_end_dates
    # ========================================================================
#    print("=" * 60)
#    print("DEPENDENCY RESOLUTION: FIRST PASS (Building end dates map)")
#    print("=" * 60)
    
    for maquina in maquinas:
        machine_id = maquina.id_maquina
        
        # 1. Fetch "Native" Tasks from SQL (Tasks physically assigned to this machine)
        # --------------------------------------------------------------------------------
        filtros = request.GET.copy()
        
        machine_filter = {'machine_ids': [machine_id]}
        
        # Merge URL/Session specific projects
        if proyectos_val:
             machine_filter['proyectos'] = [p.strip() for p in proyectos_val.split(',') if p.strip()]
             
        if planned_ids:
            machine_filter['id_orden_in'] = planned_ids
        elif request.GET.get('id_orden'):
            machine_filter['id_orden'] = request.GET.get('id_orden')

        if search_active:
            native_tasks = get_planificacion_data(machine_filter) 
        else:
            native_tasks = []
        
        # 2. Filter OUT tasks that were virtually moved AWAY
        # --------------------------------------------------------------------------------
        active_tasks = []
        
        # Prepare current machine identifiers
        current_machine_code = str(machine_id).strip()
        current_machine_name = str(maquina.nombre).strip()

        for t in native_tasks:
            try:
                oid = int(t.get('Idorden', 0))
            except (ValueError, TypeError):
                oid = 0
            
            # Check Virtual Map
            if oid in virtual_overrides:
                override_data = virtual_overrides[oid]
                target_machine = str(override_data['maquina']).strip()
                
                # Check if the target machine matches THIS machine (either by Code or Name)
                # If it matches, we keep it (it was 'moved' here, or stayed here).
                # If it doesn't match, it was moved AWAY.
                if target_machine == current_machine_code or target_machine == current_machine_name:
                    # Is it hidden?
                    if oid not in hidden_ids:
                        active_tasks.append(t)
            else:
                 # No virtual move. Keep it unless hidden.
                 if oid not in hidden_ids:
                     active_tasks.append(t)
                 
        # 3. Add tasks that were virtually moved IN (from other machines)
        # --------------------------------------------------------------------------------
        # keys in tasks_moved_in_map could be Codes OR Names. Check both.
        moved_in_ids = []
        if current_machine_code in tasks_moved_in_map:
            moved_in_ids.extend(tasks_moved_in_map[current_machine_code])
        if current_machine_name in tasks_moved_in_map:
             # Avoid duplicates if Code == Name or overlap
             new_ids = tasks_moved_in_map[current_machine_name]
             moved_in_ids.extend([i for i in new_ids if i not in moved_in_ids])
        
        if moved_in_ids:
            inbound_filter = {}
            if proyectos_val:
                 inbound_filter['proyectos'] = machine_filter.get('proyectos', [])

            if planned_ids:
                inbound_filter['id_orden_in'] = moved_in_ids
            else:
                inbound_filter['id_orden_in'] = moved_in_ids
            
            # Since moved_in_ids means we have data
            extra_tasks = get_planificacion_data(inbound_filter)
            
            # Merge unique tasks (avoid duplicates if native query somehow caught them)
            existing_ids = set(t['Idorden'] for t in active_tasks)
            for t in extra_tasks:
                t_id = t['Idorden']
                if t_id not in existing_ids and t_id not in hidden_ids:
                    active_tasks.append(t)

        
        # Deduplicate tasks by Idorden just in case
        unique_tasks_map = {}
        for t in active_tasks:
            # Use string key for robust deduplication
            tid = str(t.get('Idorden'))
            if tid not in unique_tasks_map:
                unique_tasks_map[tid] = t
        
        tasks = list(unique_tasks_map.values())
        
        # --- Apply Visual Priority Sorting AND Manual Time Override ---
        for idx, item in enumerate(tasks):
             # Default Priority (preserve SQL order)
             default_prio = (idx + 1) * 1000.0
             
             p_id = item['Idorden']
             
             # Robust Lookup: Try raw, then int, then str
             override_found = False
             ov_data = None
             
             if p_id in virtual_overrides:
                 ov_data = virtual_overrides[p_id]
                 override_found = True
             else:
                 try:
                     p_id_int = int(p_id)
                     if p_id_int in virtual_overrides:
                         ov_data = virtual_overrides[p_id_int]
                         override_found = True
                 except (ValueError, TypeError):
                     pass
             
             if override_found and ov_data:
                 item['OrdenVisual'] = float(ov_data['prioridad'])
                 
                 # Apply Time Override
                 if ov_data.get('tiempo_manual') is not None:
                     item['Tiempo_Proceso'] = float(ov_data['tiempo_manual'])
                     item['CalculadoManual'] = True

                 if ov_data.get('nivel_manual') is not None:
                      pass
             else:
                 item['OrdenVisual'] = default_prio
                 
        # Jerarquía estricta en cascada:
        #   1. Prioridad Proyecto  ASC  → menor número = mayor prioridad.
        #   2. Prioridad Artículo  ASC  → menor número = mayor prioridad.
        #   3. Nivel Planificación  DESC → mayor nivel va antes (negativo).
        #   4. Idorden (OP)         ASC  → desempate cuando los niveles son idénticos.
        #   5. secuencia_proceso / OrdenSecuencia / OrdenVisual  → desempates de seguridad.
        tasks.sort(key=lambda x: (
            int(x.get('prioridad_proyecto', 999)),
            int(x.get('prioridad_pieza', 9999)),
            -int(x.get('Nivel_Planificacion', 0) or 0),
            int(x.get('Idorden', 9999999)),
            x.get('secuencia_proceso', 999),
            x.get('OrdenSecuencia', 999999),
            x.get('OrdenVisual', 999999)
        ))
        
        # Re-normalize priorities within the final resulting order
        for idx, item in enumerate(tasks):
            item['OrdenVisual'] = (idx + 1) * 1000
        
        # Store for second pass
        machine_tasks_map[machine_id] = {
            'maquina': maquina,
            'tasks': tasks
        }
        
        # FIRST PASS: Calculate WITHOUT dependency constraints
        # This builds the initial end_dates map
        calculated_tasks = calculate_timeline(maquina, tasks, start_date=start_simulation, task_min_start_times=None)
        
        # Update Global End Dates with FIRST PASS results
        for ct in calculated_tasks:
             ct_id = ct.get('Idorden')
             ct_end = ct.get('end_date')
             if ct_id and ct_end:
                 if ct_id not in global_task_end_dates or ct_end > global_task_end_dates[ct_id]:
                     global_task_end_dates[ct_id] = ct_end
                     
                     # Debug logging for our specific tasks
                     if ct_id in [46762, 46759]:
                         print(f"  [END DATE RECORDED] Task {ct_id} ends at {ct_end}")
        
        print(f"  Machine {machine_id}: Calculated {len(calculated_tasks)} tasks")

    # ========================================================================
    # SECOND PASS: Recalculate ONLY tasks with dependencies (MULTI-PASS)
    # ========================================================================
#    print("\n" + "=" * 60)
#    print("DEPENDENCY RESOLUTION: SECOND PASS (Applying dependencies - Multi-Pass)")
#    print("=" * 60)
    
    # Identify which tasks have dependencies
    tasks_with_dependencies = set(dependency_map.keys())
    
    # We will store the FINAL result for each machine here
    final_timeline_map = {} 

    # We run this loop multiple times to propagate dependency changes across machines.
    # e.g. Machine A changes -> affects Machine B -> affects Machine A's later tasks.
    NUM_PASSES = 3
    
    for pass_idx in range(NUM_PASSES):
        print(f"\n--- Resolution Pass {pass_idx + 1}/{NUM_PASSES} ---")
        changes_detected = False # We could optimize to stop if no changes, but fixed passes is safer/simpler
        
        for machine_id, machine_data in machine_tasks_map.items():
            maquina = machine_data['maquina']
            tasks = machine_data['tasks']
            
            # Check if ANY task in this machine has dependencies OR if we just want to update strictly
            # Actually, even if *this* machine has no dependencies, its tasks might be NEEDED by others.
            # So if we are in Pass 1, we might rely on Pass 0 (Simulated).
            # But 'calculate_timeline' is deterministic if inputs (min_start_times) don't change.
            
            min_start_times = {}
            has_deps_here = False
            
            for t in tasks:
                t_id = t.get('Idorden')
                if t_id in dependency_map:
                    has_deps_here = True
                    preds = dependency_map[t_id]
                    max_pred_end = None
                    
                    for pid in preds:
                        if pid in global_task_end_dates:
                            end_date = global_task_end_dates[pid]
                            if max_pred_end is None or end_date > max_pred_end:
                                max_pred_end = end_date
                    
                    if max_pred_end:
                        min_start_times[t_id] = max_pred_end
                        
                        # Debug logging for our specific tasks
                        if t_id in [46762, 46759]:
                            print(f"  [DEPENDENCY APPLIED] Task {t_id} must start after {max_pred_end} (from predecessors: {preds})")
            
            # Optimization: If no dependencies here, and we already calculated in Pass 0 (First Pass), 
            # we technically don't need to re-run unless we want to be super safe. 
            # But First Pass didn't use `min_start_times`. So YES, we must run at least once if there are deps.
            # If Pass > 0 and no input changes... but let's just run it. It's fast.
            
            recalculated_tasks = calculate_timeline(maquina, tasks, start_date=start_simulation, task_min_start_times=min_start_times)
            
            # SAVE RESULT
            final_timeline_map[machine_id] = {
                'machine': maquina,
                'tasks': recalculated_tasks
            }
            
            # CRITICAL: Update Global End Dates LIVE for next machines/next pass
            for ct in recalculated_tasks:
                 ct_id = ct.get('Idorden')
                 ct_end = ct.get('end_date')
                 if ct_id and ct_end:
                     global_task_end_dates[ct_id] = ct_end
            
            print(f"  Machine {machine_id}: Recalculated {len(recalculated_tasks)} tasks")

    # Build final list
    # Ensure preservation of order if possible (not strictly required since we group by machine)
    from .planning_service import get_active_maintenances
    for machine_id in machine_tasks_map.keys(): # Use original keys ordering
        if machine_id in final_timeline_map:
            row = final_timeline_map[machine_id]
            m = row['machine']
            if hasattr(m, 'id_maquina') and m.id_maquina != 'MAC00':
                row['maintenances'] = get_active_maintenances(m)
            else:
                row['maintenances'] = []
            timeline_data.append(row)
    
    # FILTER: REMOVED per user request ("no las ocultes")
    # We keep empty rows to allow Drag and Drop to empty machines
    # timeline_data = [row for row in timeline_data if row['tasks']]

#    print("=" * 60)
#    print(f"DEPENDENCY RESOLUTION COMPLETE")
#    print(f"Total tasks processed: {len(global_task_end_dates)}")
#    print("=" * 60 + "\n")
        
    # 3. Determine Visual Bounds (Min/Max working hours)
    global_min_h = 24
    global_max_h = 0
    has_schedules = False
    
    for m in maquinas:
        for h in m.horarios.all():
            has_schedules = True
            if h.hora_inicio.hour < global_min_h:
                global_min_h = h.hora_inicio.hour
            if h.hora_fin.hour > global_max_h:
                global_max_h = h.hora_fin.hour
    
    if not has_schedules:
        global_min_h = 7
        global_max_h = 18
    else:
        # Buffer ensures we see the closing hour block?
        # If max is 22, range(7, 22) stops at 21:59. Correct.
        pass
        
    if global_max_h <= global_min_h:
        global_max_h = 23
        global_min_h = 0

    # 4. Generate Time Columns (Filtering non-working hours)
    # Align start to the min_hour of the start day
    min_date = start_simulation.replace(hour=global_min_h, minute=0, second=0, microsecond=0)
    
    # Calculate Max Date from tasks
    calc_max_date = min_date + timedelta(hours=48)
    for row in timeline_data:
        for t in row['tasks']:
            if t['end_date'] and t['end_date'] > calc_max_date:
                calc_max_date = t['end_date']
    
    # Determine which days are "Working Days" to display
    # Iterate from min_date to calc_max_date
    # Rules: 
    # - Always include Mon-Fri (LV)
    # - Include Sat (SA) ONLY if at least one machine has simple schedule 'SA' or we default to showing it?
    #   Let's check if any machine has 'SA' in its schedules.
    show_saturdays = False
    for m in maquinas:
        for h in m.horarios.all():
            if h.dia == 'SA':
                show_saturdays = True
                break
        if show_saturdays: break
        
    # Strict list of valid dates for columns
    valid_dates = []
    day_pointer = min_date.date()
    end_date_limit = calc_max_date.date()
    day_count = (end_date_limit - day_pointer).days + 5 # Buffer
    
    for d in range(day_count):
        current_day = day_pointer + timedelta(days=d)
        wd = current_day.weekday() # 0=Mon, 6=Sun
        
        is_working_day = False
        if 0 <= wd <= 4: # Mon-Fri
            is_working_day = True
        elif wd == 5 and show_saturdays: # Sat
            is_working_day = True
            
        if is_working_day:
            valid_dates.append(current_day)
            
    # Map Date -> Column Index Start (Visual Day Index)
    # e.g. Mon=0, Tue=1, (Sat skip), Mon=2...
    date_to_visual_index = { d: i for i, d in enumerate(valid_dates) }

    # Generate Columns
    time_columns = []
    slots_per_day = global_max_h - global_min_h
    
    for d in valid_dates:
        for h in range(global_min_h, global_max_h):
             dt = datetime.combine(d, datetime.min.time()) + timedelta(hours=h)
             time_columns.append(dt)

    # =========================================================
    # 5. POSICIONAMIENTO DEFINITIVO - ANTI-SOLAPAMIENTO
    # Regla: visual_left[n] = max(time_pos[n], cursor[n-1] + GAP)
    # MIN_WIDTH = 100px universal — ninguna card puede ser menor.
    # =========================================================
    COL_WIDTH  = 100   # px por hora
    MIN_WIDTH  = 40    # px mínimo de card
    SAFETY_GAP = 6     # px de aire entre cards
    
    def _time_to_px(dt_obj):
        day_idx = date_to_visual_index.get(dt_obj.date(), 0)
        h_diff  = (dt_obj.hour - global_min_h) + (dt_obj.minute / 60.0)
        if h_diff < 0: h_diff = 0
        return ((day_idx * slots_per_day) + h_diff) * COL_WIDTH

    for row in timeline_data:

        # Paso 1 — lista unificada (tareas + mantenimientos) con ancho_total_elemento
        events = []
        for t in row['tasks']:
            if not t.get('start_date'):
                continue
            raw_left        = _time_to_px(t['start_date'])
            duration_px     = t['duration_real'] * COL_WIDTH
            # ancho_total_elemento: mínimo 100px siempre
            ancho_total     = max(MIN_WIDTH, duration_px)
            events.append({'obj': t, 'raw_left': raw_left, 'ancho': ancho_total, 'is_maint': False})

        for m in row.get('maintenances', []):
            raw_left    = _time_to_px(m['start'])
            maint_px    = (m['end'] - m['start']).total_seconds() / 3600.0 * COL_WIDTH
            ancho_total = max(MIN_WIDTH, maint_px)
            events.append({'obj': m, 'raw_left': raw_left, 'ancho': ancho_total, 'is_maint': True})

        # Paso 2 — ordenar por tiempo real
        events.sort(key=lambda e: e['raw_left'])

        # Paso 3 — push acumulativo
        # cursor  = borde derecho del último elemento YA posicionado
        cursor = -9999.0
        for ev in events:
            # REGLA CENTRAL: left[n] = max(time_pos, fin_anterior + gap)
            final_left = max(ev['raw_left'], cursor + SAFETY_GAP)
            ev['obj']['visual_left']  = round(final_left, 2)
            ev['obj']['visual_width'] = round(ev['ancho'],  2)
            cursor = final_left + ev['ancho']
            # Diagnóstico en consola
            oid  = ev['obj'].get('Idorden', 'MAINT')
            is_d = ev['obj'].get('is_delayed', False)
            print(f"  [POS] OP={oid} delayed={is_d} left={final_left:.0f} width={ev['ancho']:.0f} cursor={cursor:.0f}")



    # Build dependencies list for JSON (for visualization)
    dependencies_list = []
    for succ_id, pred_ids in dependency_map.items():
        for pred_id in pred_ids:
            dependencies_list.append({'pred': pred_id, 'succ': succ_id})
    
    context = {
        'timeline_data': timeline_data,
        'time_columns': time_columns,
        'start_date': min_date,
        'dependencies_json': json.dumps(dependencies_list),
    }
    return render(request, 'produccion/planificacion_visual.html', context)

def maquina_config_create_update(request, pk=None):
    """
    Create or Update a machine.
    """
    maquina = None
    if pk:
        maquina = get_object_or_404(MaquinaConfig.objects.using('default'), pk=pk)
    
    if request.method == 'POST':
        id_maquina = request.POST.get('id_maquina')
        nombre = request.POST.get('nombre')
        
        if not id_maquina or not nombre:
            messages.error(request, "Todos los campos son obligatorios")
            return redirect('maquina_config_list')
            
        try:
            if maquina:
                 # Update
                 new_id = id_maquina # From POST
                 
                 if new_id != maquina.pk:
                     # ID Changed: Rename Logic
                     if MaquinaConfig.objects.using('default').filter(pk=new_id).exists():
                         messages.error(request, f"El ID '{new_id}' ya existe. No se puede renombrar.")
                         return render(request, 'produccion/maquina_config_form.html', {'maquina': maquina})
                     
                     # 1. Create New
                     new_maquina = MaquinaConfig.objects.using('default').create(id_maquina=new_id, nombre=nombre)
                     
                     # 2. Move Related Horarios
                     for horario in maquina.horarios.all():
                         horario.maquina = new_maquina
                         horario.save(using='default')
                         
                     # 3. Delete Old
                     maquina.delete(using='default')
                     messages.success(request, f"MÃ¡quina renombrada a '{new_id}' y actualizada correctamente")
                 else:
                     # Standard Update
                     maquina.nombre = nombre
                     maquina.save(using='default')
                     messages.success(request, "MÃ¡quina actualizada correctamente")
            else:
                 # Create
                 if MaquinaConfig.objects.using('default').filter(pk=id_maquina).exists():
                     messages.error(request, "El ID de mÃ¡quina ya existe")
                     return redirect('maquina_config_list')
                     
                 MaquinaConfig.objects.using('default').create(id_maquina=id_maquina, nombre=nombre)
                 messages.success(request, "MÃ¡quina creada correctamente")
                 
            return redirect('maquina_config_list')
        except Exception as e:
            messages.error(request, f"Error: {e}")
    
    return render(request, 'produccion/maquina_config_form.html', {'maquina': maquina})

def maquina_config_delete(request, pk):
    if request.method == 'POST':
        maquina = get_object_or_404(MaquinaConfig.objects.using('default'), pk=pk)
        maquina_name = maquina.nombre
        maquina.delete(using='default')
        
        # Signal frontend to CLEAR visual Gantt state
        request.session['gantt_needs_clear'] = True
        messages.success(request, f"Máquina {maquina_name} eliminada. Gantt limpiado.")
        
    return redirect('maquina_config_list')

def horario_maquina_create(request, maquina_id):
    if request.method == 'POST':
        maquina = get_object_or_404(MaquinaConfig.objects.using('default'), pk=maquina_id)
        dia = request.POST.get('dia')
        hora_inicio = request.POST.get('hora_inicio')
        hora_fin = request.POST.get('hora_fin')
        
        try:
            HorarioMaquina.objects.using('default').create(
                maquina=maquina,
                dia=dia,
                hora_inicio=hora_inicio,
                hora_fin=hora_fin
            )
            messages.success(request, "Horario agregado")
        except Exception as e:
            messages.error(request, f"Error al agregar horario: {e}")
            
    return redirect('maquina_config_list')

def horario_maquina_delete(request, pk):
    if request.method == 'POST':
        horario = get_object_or_404(HorarioMaquina.objects.using('default'), pk=pk)
        horario.delete(using='default')
        messages.success(request, "Horario eliminado")
    return redirect('maquina_config_list')

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import timedelta



# --- Feriados Views ---

from .models import Feriado
from .forms import FeriadoForm, MantenimientoMaquinaForm
from django.db.models import Q
from .models import MantenimientoMaquina

# ==========================================
# GESTIÓN DE MANTENIMIENTOS
# ==========================================

def mantenimiento_list(request):
    mantenimientos = MantenimientoMaquina.objects.using('default').all().order_by('-fecha_inicio')
    
    # Optional filtering
    maq_id = request.GET.get('maquina')
    if maq_id:
        mantenimientos = mantenimientos.filter(maquina_id=maq_id)
        
    estado = request.GET.get('estado')
    if estado:
        mantenimientos = mantenimientos.filter(estado=estado)
        
    context = {
        'mantenimientos': mantenimientos,
        'maquinas': MaquinaConfig.objects.using('default').all(),
        'selected_maq': maq_id,
        'selected_estado': estado
    }
    return render(request, 'produccion/mantenimiento_list.html', context)

def mantenimiento_create_update(request, pk=None):
    if pk:
        mantenimiento = get_object_or_404(MantenimientoMaquina.objects.using('default'), pk=pk)
        title = "Editar Mantenimiento"
    else:
        mantenimiento = None
        title = "Programar Mantenimiento"
        
    if request.method == 'POST':
        form = MantenimientoMaquinaForm(request.POST, instance=mantenimiento)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.save(using='default')
            messages.success(request, f'Mantenimiento {"actualizado" if pk else "programado"} exitosamente.')
            return redirect('mantenimiento_list')
    else:
        form = MantenimientoMaquinaForm(instance=mantenimiento)
        
    context = {
        'form': form,
        'title': title,
        'is_edit': bool(pk)
    }
    return render(request, 'produccion/mantenimiento_form.html', context)

def mantenimiento_delete(request, pk):
    mantenimiento = get_object_or_404(MantenimientoMaquina.objects.using('default'), pk=pk)
    if request.method == 'POST':
        mantenimiento.delete()
        messages.success(request, 'Mantenimiento eliminado correctamente.')
        return redirect('mantenimiento_list')
    return render(request, 'produccion/mantenimiento_confirm_delete.html', {'mantenimiento': mantenimiento})


# ==========================================
# GESTIÓN DE FERIADOS
# ==========================================

def feriado_list(request):
    """
    Lista todos los feriados con filtros opcionales y paginación.
    """
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # Filtros
    year_filter = request.GET.get('year')
    
    feriados = Feriado.objects.filter(activo=True)
    
    # Filtro por año
    selected_year = None
    if year_filter:
        try:
            selected_year = int(year_filter)
            feriados = feriados.filter(fecha__year=selected_year)
        except ValueError:
            pass
    
    # Obtener años disponibles para el filtro
    years = Feriado.objects.dates('fecha', 'year', order='DESC')
    available_years = [d.year for d in years]
    
    # Paginación
    paginator = Paginator(feriados, 8)  # 8 feriados por página
    page = request.GET.get('page')
    
    try:
        feriados_page = paginator.page(page)
    except PageNotAnInteger:
        # Si page no es un entero, mostrar la primera página
        feriados_page = paginator.page(1)
    except EmptyPage:
        # Si page está fuera de rango, mostrar la última página
        feriados_page = paginator.page(paginator.num_pages)
    
    context = {
        'feriados': feriados_page,
        'available_years': available_years,
        'selected_year': selected_year,
    }
    
    return render(request, 'produccion/feriado_list.html', context)


def feriado_create(request):
    """
    Crear un nuevo feriado desde el formulario simplificado.
    """
    if request.method == 'POST':
        # Obtener datos del formulario simplificado
        fecha = request.POST.get('fecha')
        descripcion = request.POST.get('descripcion')
        
        if not fecha or not descripcion:
            messages.error(request, 'Por favor complete todos los campos.')
            return redirect('feriado_list')
        
        try:
            # Verificar que no exista otro feriado en la misma fecha
            if Feriado.objects.filter(fecha=fecha).exists():
                messages.error(request, f'Ya existe un feriado registrado para la fecha {fecha}')
                return redirect('feriado_list')
            
            # Crear el feriado con valores por defecto
            feriado = Feriado.objects.create(
                fecha=fecha,
                descripcion=descripcion,
                tipo_jornada='NO',  # Por defecto no se trabaja
                activo=True
            )
            messages.success(request, f'Feriado "{feriado.descripcion}" creado exitosamente.')
            return redirect('feriado_list')
        except Exception as e:
            messages.error(request, f'Error al crear el feriado: {str(e)}')
            return redirect('feriado_list')
    else:
        form = FeriadoForm()
    
    return render(request, 'produccion/feriado_form.html', {
        'form': form,
        'title': 'Crear Nuevo Feriado',
        'button_text': 'Crear Feriado'
    })


def feriado_update(request, pk):
    """
    Editar un feriado existente.
    """
    feriado = get_object_or_404(Feriado, pk=pk)
    
    if request.method == 'POST':
        form = FeriadoForm(request.POST, instance=feriado)
        if form.is_valid():
            feriado = form.save()
            messages.success(request, f'Feriado "{feriado.descripcion}" actualizado exitosamente.')
            return redirect('feriado_list')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = FeriadoForm(instance=feriado)
    
    return render(request, 'produccion/feriado_form.html', {
        'form': form,
        'feriado': feriado,
        'title': f'Editar Feriado: {feriado.descripcion}',
        'button_text': 'Guardar Cambios'
    })


def feriado_delete(request, pk):
    """
    Eliminar un feriado.
    """
    feriado = get_object_or_404(Feriado, pk=pk)
    
    if request.method == 'POST':
        descripcion = feriado.descripcion
        feriado.delete()
        messages.success(request, f'Feriado "{descripcion}" eliminado exitosamente.')
        return redirect('feriado_list')
    
    return render(request, 'produccion/feriado_confirm_delete.html', {
        'feriado': feriado
    })


@csrf_exempt
def feriado_toggle_planifica(request, pk):
    """
    API para cambiar rÃ¡pidamente el estado de planificaciÃ³n de un feriado.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        feriado = get_object_or_404(Feriado, pk=pk)
        feriado.se_planifica = not feriado.se_planifica
        feriado.save()
        
        return JsonResponse({
            'status': 'ok',
            'se_planifica': feriado.se_planifica,
            'message': f'Feriado {"se trabajarÃ¡" if feriado.se_planifica else "no se trabajarÃ¡"}'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def feriado_toggle_activo(request, pk):
    """
    API para activar/desactivar un feriado.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        feriado = get_object_or_404(Feriado, pk=pk)
        feriado.activo = not feriado.activo
        feriado.save()
        
        return JsonResponse({
            'status': 'ok',
            'activo': feriado.activo,
            'message': f'Feriado {"activado" if feriado.activo else "desactivado"}'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def feriado_update_jornada(request, pk):
    """
    API para actualizar el tipo de jornada de un feriado.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        feriado = get_object_or_404(Feriado, pk=pk)
        
        # Obtener el nuevo tipo de jornada del body
        data = json.loads(request.body)
        nuevo_tipo = data.get('tipo_jornada', 'NO')
        
        # Validar que el tipo sea válido
        if nuevo_tipo not in ['NO', 'MEDIO', 'SI']:
            return JsonResponse({'error': 'Tipo de jornada inválido'}, status=400)
        
        # Actualizar el feriado
        feriado.tipo_jornada = nuevo_tipo
        feriado.save()
        
        return JsonResponse({
            'status': 'ok',
            'tipo_jornada': feriado.tipo_jornada,
            'message': f'Tipo de jornada actualizado a {feriado.get_tipo_jornada_display()}'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =========================================================================
# NEW SHARED LOGIC IMPLEMENTATION
# =========================================================================

def planificacion_visual(request):
    """
    Visual Gantt Chart View.
    Uses shared logic from gantt_logic.py
    """
    from django.utils import timezone
    from .models import Scenario
    
    scenario_id_param = request.GET.get('scenario_id', None)
    
    if scenario_id_param and scenario_id_param != 'null':
        try:
            scenario_id = int(scenario_id_param)
        except ValueError:
            scenario_id = None
    else:
        scenario_id = None
        
    if not scenario_id:
        active_scenario = get_active_scenario(request)
        if active_scenario:
            params = request.GET.copy()
            # Quitar el posible 'null' en string sucio
            if 'scenario_id' in params:
                del params['scenario_id']
            params['scenario_id'] = active_scenario.id
            return redirect(f"{request.path}?{params.urlencode()}")
            
    active_scenario = get_active_scenario(request)

    import urllib.parse
    all_scenarios = Scenario.objects.using('default').all().order_by('-es_principal', 'nombre')
    
    proyectos_value = request.GET.get('proyectos', '')
    if proyectos_value:
        proyectos_value = urllib.parse.unquote(proyectos_value).replace('%2C', ',').strip()
    
    if not proyectos_value and active_scenario:
        from .models import PlannedTask
        # SNAPSHOT RULE: Primero leer proyectos de las tareas ya confirmadas al Gantt.
        # Si no hay ninguna confirmada, fallback a todas las PlannedTasks del escenario.
        db_proyectos = list(PlannedTask.objects.using('default').filter(
            scenario=active_scenario,
            enviado_a_gantt=True
        ).values_list('proyecto_code', flat=True).distinct())
        if not db_proyectos:
            # FALLBACK: No hay snapshot congelado → usar todas las tareas del escenario
            db_proyectos = list(PlannedTask.objects.using('default').filter(
                scenario=active_scenario
            ).values_list('proyecto_code', flat=True).distinct())
        if db_proyectos:
            proyectos_value = ','.join(p for p in db_proyectos if p)
            
    # ---- REEMPLAZÁ DESDE ACÁ ----
    # 1. Capturamos el modo desde el GET o desde el POST
    plan_mode = request.GET.get('plan_mode') or request.POST.get('plan_mode')

    # 2. Si no vino en la petición actual, lo recuperamos de la sesión
    if not plan_mode:
        plan_mode = request.session.get('last_plan_mode', 'manual')

    # 3. Guardamos el modo actual en la sesión para que lo recuerde al actualizar
    request.session['last_plan_mode'] = plan_mode
    # ---- HASTA ACÁ ----

    # Check if manual render is triggered
    graficar = request.GET.get('graficar') == '1'
    
    if not graficar:
        # Fast load, empty Gantt view
        context = {
            'timeline_data': [],
            'time_columns': [],
            'start_date': timezone.now(),
            'dependencies_json': json.dumps([]),
            'today': timezone.now(),
            'total_width': 0,
            'system_alerts': [],
            'analysis': {'machines': [], 'project_alerts': []},
            'all_scenarios': all_scenarios,
            'active_scenario': active_scenario,
            'plan_mode': plan_mode,  # Esto se mantiene igual, ya usa la variable nueva
            'gantt_needs_clear': False,
            'any_rendering_capped': False,
            'proyectos_value': proyectos_value,
        }
        return render(request, 'produccion/planificacion_visual.html', context)
        
    # Use shared logic
    data = get_gantt_data(request)
    
    # Extract data for context
    timeline_data = data['timeline_data']
    
    # --- 4. RENDERING OPTIMIZATION (OpenCode Fix) ---
    # User requested: "Limitá la cantidad de bloques que el frontend intenta dibujar simultáneamente si el sistema detecta una inconsistencia masiva de fechas."
    MAX_BLOCKS_PER_MACHINE = 150 # Safety cap
    total_blocks = 0
    for row in timeline_data:
        if len(row['tasks']) > MAX_BLOCKS_PER_MACHINE:
            print(f"DEBUG: [Optimization] Capping tasks for machine {row['machine'].id_maquina} from {len(row['tasks'])} to {MAX_BLOCKS_PER_MACHINE}")
            row['tasks'] = row['tasks'][:MAX_BLOCKS_PER_MACHINE]
            row['rendering_capped'] = True
        total_blocks += len(row['tasks'])
    
    if total_blocks > 2000: # Total global cap
        print(f"DEBUG: [Optimization] Massive block count detected ({total_blocks}). UI performance might be degraded.")
    time_columns = data['time_columns']
    valid_dates = data['valid_dates']
    start_simulation = data['start_simulation']
    dependency_map = data['dependency_map']
    global_min_h = data['global_min_h']
    global_max_h = data['global_max_h']
    day_max_hours = data.get('day_max_hours', {})
    date_start_col = data.get('date_start_col', {})

    # =========================================================
    # 5. POSICIONAMIENTO DINÁMICO (ANTI-SOLAPAMIENTO)
    # Regla: visual_left = max(posición_por_tiempo, fin_anterior + 6px)
    # Para OPs con atraso, garantizamos un ancho visual de 100px.
    # =========================================================
    COL_WIDTH = 100  # px por hora — debe coincidir con .time-hour { width: 100px } en el CSS
    DAY_GAP = 10
    
    # Mapa de fechas a índice visual
    date_to_day_idx = {d: i for i, d in enumerate(valid_dates)}
    
    for row in timeline_data:
        # --- Paso 1: Recolectar todos los elementos de la fila (ordenados por tiempo) ---
        elements = []
        for t in row['tasks']:
            t_start = t.get('start_date')
            if not t_start: continue
            
            day_idx = date_to_day_idx.get(t_start.date(), 0)
            day_col_start = date_start_col.get(t_start.date(), 0)
            hour_diff = float((t_start.hour - global_min_h) + (t_start.minute / 60.0))
            if hour_diff < 0: hour_diff = 0
            
            time_left = (day_col_start + hour_diff) * COL_WIDTH + (day_idx * DAY_GAP)
            duration_px = t.get('duration_real', 0) * COL_WIDTH
            if duration_px <= 0: continue
            elements.append({'obj': t, 'time_left': time_left, 'duration_px': duration_px, 'is_maint': False})

        for m in row.get('maintenances', []):
            m_s = m.get('start')
            if not m_s: continue
            m_dur_px = (m['end'] - m['start']).total_seconds() / 3600.0 * COL_WIDTH
            day_idx = date_to_day_idx.get(m_s.date(), 0)
            day_col_start = date_start_col.get(m_s.date(), 0)
            hour_diff = float((m_s.hour - global_min_h) + (m_s.minute / 60.0))
            if hour_diff < 0: hour_diff = 0
            time_left = (day_col_start + hour_diff) * COL_WIDTH + (day_idx * DAY_GAP)
            elements.append({'obj': m, 'time_left': time_left, 'duration_px': m_dur_px, 'is_maint': True})

        # --- No ordenamos 'elements' --- 
        # Al no ordenar, respetamos EXACTAMENTE el orden en que las tareas vinieron de la tabla
        # (que a su vez está dictado por el backend y OrdenVisual original).

        # --- Paso 2: Posicionamiento Acumulativo Estricto (Cascada) ---
        cursor_card_end = 0.0   # La fila arranca en 0px
        cursor_badge_end = -9999.0
        stagger_level = 0
        
        for el in elements:
            obj = el['obj']
            try:
                obj['Idorden'] = str(int(float(obj.get('Idorden', 0))))
            except:
                obj['Idorden'] = str(obj.get('Idorden', ''))

            # Duración EXACTA de la tabla: reflejo directo de Tiempo_Proceso.
            # Sin ancho mínimo, sin márgenes, sin optimizaciones que alteren la escala temporal original.
            duration_px = el['duration_px']

            # Posicionamiento exacto basado en el motor de planificación
            final_left = el['time_left']
            cursor_card_end = final_left + duration_px

            obj['visual_left']  = round(final_left, 2)
            obj['visual_width'] = round(duration_px, 2)
            
            # 3. Lógica de Escalonamiento de Badges (Staggering)
            has_badge = obj.get('is_delayed') and obj.get('segment_index', 0) == 0
            if has_badge:
                if final_left < (cursor_badge_end + 5):
                     stagger_level = (stagger_level + 1) % 2
                else:
                     stagger_level = 0
                
                obj['badge_stagger'] = stagger_level
                cursor_badge_end = final_left + 95
            
            # El cursor físico ya fue adelantado en base a duration_px
            # (No actualizamos por natural_w para evitar inyectar aire matemático)

        # Forzar el reordenamiento del diccionario original para que el HTML Render ({% for task in row.tasks %})
        # recorra las tareas y mantenimientos exactamente en nuestra subsecuencia correcta sumada
        row['tasks'] = [el['obj'] for el in elements if not el['is_maint']]
        row['maintenances'] = [el['obj'] for el in elements if el['is_maint']]


    # Build time columns with gap info
    time_columns_data = []
    last_date = None
    total_gaps = 0
    
    # Handle empty or non-datetime time_columns (e.g., range object from early return)
    if not time_columns or not isinstance(time_columns[0] if time_columns else None, datetime):
        # Return empty structure
        time_columns_data = [{'datetime': None, 'is_day_start': True}]
    else:
        for dt in time_columns:
            curr_date = dt.date()
            is_day_start = (curr_date != last_date)
            if is_day_start and last_date is not None:
                total_gaps += 1
            
            time_columns_data.append({
                'datetime': dt,
                'is_day_start': is_day_start
            })
            last_date = curr_date

    # Build dependencies list for JSON
    # To handle hidden tasks, we need to find the "first visible predecessor" for each visible task
    dependencies_list = []
    visible_tids = {str(t['Idorden']) for row in timeline_data for t in row['tasks']}
    
    def get_visible_preds(tid, visited=None):
        if visited is None: visited = set()
        if tid in visited: return []
        visited.add(tid)
        
        preds = dependency_map.get(tid, [])
        v_preds = []
        for pid in preds:
            s_pid = str(pid)
            # Try cleaning for match
            try: clean_p = str(int(float(s_pid)))
            except: clean_p = s_pid
            
            if clean_p in visible_tids:
                v_preds.append(clean_p)
            else:
                # Recursively look for visible predecessors of this hidden task
                v_preds.extend(get_visible_preds(s_pid, visited))
        return v_preds

    for succ_id in visible_tids:
        v_preds = get_visible_preds(succ_id)
        for pred_id in set(v_preds): # Deduplicate
            dependencies_list.append({'pred': pred_id, 'succ': succ_id})
    
    print(f"DEBUG: [Dependencies] Grafo reconstruido (saltando ocultos): {len(dependencies_list)} vínculos encontrados.")
    
    print(f"DEBUG: [Dependencies] Grafo generado: {len(dependencies_list)} vínculos encontrados.")

    # =========================================================
    # BUG FIX: Generar LINKS por PIEZA recorriendo las tareas directamente.
    # El dependency_map puede estar incompleto para proyectos recién cargados
    # (los niveles del ERP no se reconocieron), por lo que generamos un grafo
    # alternativo agrupando por (ProyectoCode, Mstnmbr) y uniendo operaciones
    # consecutivas en orden DESCENDENTE de nivel_planificacion.
    # Formato: [{id, source, target, type}, ...] — type "0" = Finish-to-Start
    # =========================================================
    links_list = []
    piece_groups = {}
    for row in timeline_data:
        for t in row.get('tasks', []):
            proj = t.get('ProyectoCode') or ''
            mst = t.get('Mstnmbr') or t.get('Articulo') or 0
            try:
                oid = str(int(float(t.get('Idorden') or 0)))
            except (ValueError, TypeError):
                continue
            if not oid or oid == '0':
                continue
            nivel = t.get('nivel_planificacion')
            if nivel is None:
                nivel = t.get('Nivel_Planificacion') or t.get('Nivel') or 0
            try:
                nivel = float(nivel)
            except (ValueError, TypeError):
                nivel = 0.0
            piece_groups.setdefault((proj, mst), []).append({
                'oid': oid,
                'nivel': nivel,
            })

    link_id = 0
    for (proj, mst), ops in piece_groups.items():
        ops_sorted = sorted(ops, key=lambda x: x['nivel'], reverse=True)
        for i in range(len(ops_sorted) - 1):
            upper = ops_sorted[i]
            lower = ops_sorted[i + 1]
            if upper['nivel'] <= 0 or lower['nivel'] <= 0:
                continue
            if upper['nivel'] < lower['nivel']:
                continue
            links_list.append({
                'id': f"L{link_id}",
                'source': upper['oid'],
                'target': lower['oid'],
                'type': '0',  # Finish-to-Start
            })
            link_id += 1

    print(f"DEBUG: [Links] {len(links_list)} uniones generadas desde nivel_planificacion de las piezas.")

    # Fetch all scenarios for selector (used in template)
    from .models import Scenario
    all_scenarios = Scenario.objects.using('default').all().order_by('-es_principal', 'nombre')

    # Calculate actual total width in pixels
    calculated_total_width = (len(time_columns) * COL_WIDTH) + (total_gaps * DAY_GAP)

    context = {
        'timeline_data': timeline_data,
        'time_columns': time_columns_data, # Use the new data structure
        'start_date': start_simulation,
        'dependencies_json': json.dumps(dependencies_list),
        'links_json': json.dumps(links_list),
        'today': start_simulation,
        'total_width': calculated_total_width,
        'system_alerts': data.get('system_alerts', []),
        'analysis': data.get('analysis', {'machines': [], 'project_alerts': []}),
        'all_scenarios': all_scenarios,
        'active_scenario': data.get('active_scenario', None),
        'plan_mode': data.get('plan_mode', 'manual'),
        'gantt_needs_clear': data.get('gantt_needs_clear', False),
        'any_rendering_capped': any(row.get('rendering_capped') for row in timeline_data),
        'proyectos_value': proyectos_value,
    }


    # DEBUG: Log results for redistribution checking
    adaptive_alerts_count = len(data.get('analysis', {}).get('adaptive_alerts', []))
    print(f"DEBUG: [View] Fallas encontradas en adaptive_alerts: {adaptive_alerts_count}")

    return render(request, 'produccion/planificacion_visual.html', context)


def export_planificacion_excel(request):
    try:
        # 1. Obtener Datos
        data = get_gantt_data(request, force_run=True)
        timeline_data = data['timeline_data']
        time_columns = data['time_columns']
        global_min_h = data['global_min_h']
        global_max_h = data['global_max_h']
        active_scenario = data.get('active_scenario')
        
        if not time_columns:
             return HttpResponse("No hay datos calculados. Ejecute la planificacion visual primero.")

        # --- DYNAMIC START DATE CROP FOR EXCEL ---
        # Find the earliest start date of any task in the planning
        earliest_task_date = None
        for machine_row in timeline_data:
            for t in machine_row.get('tasks', []):
                start = t.get('start_date')
                if start:
                    task_date = start.date()
                    if earliest_task_date is None or task_date < earliest_task_date:
                        earliest_task_date = task_date

        # If we found an earliest task date, filter the time_columns to be >= earliest_task_date
        if earliest_task_date:
            time_columns = [h for h in time_columns if h.date() >= earliest_task_date]


        # 2. GENERACION EXCEL
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gantt Visual"
        ws.sheet_view.showGridLines = False 

        COLS_PER_HOUR = 6 
        hours_per_day = (global_max_h - global_min_h)
        unique_dates = []
        for dt in time_columns:
            if dt.date() not in unique_dates:
                unique_dates.append(dt.date())
        date_to_index = {d: i for i, d in enumerate(unique_dates)}
        
        # --- ESTILOS CORPORATIVOS ---
        CORP_DARK      = "27323E" # Carbon Slate
        CORP_BLUE      = "0078D4" # Microsoft Blue
        CORP_RED       = "E81123" # Microsoft Red
        CORP_BORDER    = "D2D2D2" # Light Grey
        ALIGN_CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)
        BORDER_THIN    = Border(left=Side(style='thin', color="CCCCCC"), right=Side(style='thin', color="CCCCCC"), top=Side(style='thin', color="CCCCCC"), bottom=Side(style='thin', color="CCCCCC"))
        
        # --- IDENTIFICACION DE PROYECTOS ---
        all_projs = set()
        for r in timeline_data:
            if not isinstance(r, dict) or 'tasks' not in r:
                continue
            tasks_list = r.get('tasks') or []
            for t in tasks_list:
                if isinstance(t, dict) and t.get('ProyectoCode'):
                    all_projs.add(t['ProyectoCode'])
        
        # Paleta Curada y Utilidades de Color
        PALETTE = ["0078D4", "107C10", "D83B01", "5C2D91", "008272", "A4262C", "004E8C", "498205"]
        def get_corp_color(v):
            idx = sum(ord(c) for c in str(v)) % len(PALETTE)
            return PALETTE[idx]

        def tint_color(hex_color, factor=0.85):
            """Genera una versión clara (tintada) de un color hex."""
            hex_color = hex_color.lstrip('#')
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)
            tr = int((1 - 0.15) * 255 + 0.15 * r)
            tg = int((1 - 0.15) * 255 + 0.15 * g)
            tb = int((1 - 0.15) * 255 + 0.15 * b)
            return f"{tr:02X}{tg:02X}{tb:02X}"

        def darken_color(hex_color, factor=0.7):
            """Genera una versión más oscura de un color hex."""
            hex_color = hex_color.lstrip('#')
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)
            dr = int(r * factor); dg = int(g * factor); db = int(b * factor)
            return f"{dr:02X}{dg:02X}{db:02X}"

        proyecto_color_map = {p: get_corp_color(p) for p in all_projs}

        # Helper para bordes precisos en rangos combinados
        def set_border(ws, start_row, end_row, start_col, end_col, border):
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).border = border

        from openpyxl.cell.cell import MergedCell

        # =========================================================
        # ESTÁNDAR CÁPSULA ENVOLVENTE (TARJETA CON BORDES REDONDEADOS SIMULADOS)
        # =========================================================
        def draw_floating_card(ws, label_row, task_row, start_col, end_col, color_hex, proj, op, is_critical=False, is_delayed=False, delay_days=0, is_continuation=False):
            task_text = f"PROJECT {proj}\nOP {op}" if proj != "---" else "..."
            if is_delayed and delay_days > 0:
                label_text = f"[ {delay_days}D ATRASO ]"
                label_color = "EF4444"
            elif is_critical:
                label_text = "[ RUTA CRÍTICA ]"
                label_color = "F97316"
            else:
                label_text = f"PROJECT {proj}"
                label_color = "1E3A8A"

            proj_rgb = color_hex.upper()
            bg_tint = tint_color(proj_rgb)

            # Borde grueso oscuro para simular el contorno independiente de la tarjeta
            side_thick = Side(style='medium', color="1E293B")
            
            for c in range(start_col, end_col + 1):
                # Cuerpo de Tarea
                cell_t = ws.cell(row=task_row, column=c)
                cell_t.fill = PatternFill("solid", fgColor=bg_tint)
                cell_t.border = Border(
                    left=side_thick if c == start_col else None,
                    right=side_thick if c == end_col else None,
                    top=None,
                    bottom=side_thick
                )
                if c == start_col:
                    if not isinstance(cell_t, MergedCell):
                        cell_t.value = task_text
                    cell_t.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell_t.font = Font(name='Calibri', bold=True, size=8, color="1E293B")

                # Sticker Superior
                cell_l = ws.cell(row=label_row, column=c)
                cell_l.fill = PatternFill("solid", fgColor=label_color)
                cell_l.border = Border(
                    left=side_thick if c == start_col else None,
                    right=side_thick if c == end_col else None,
                    top=side_thick,
                    bottom=None
                )
                if c == start_col:
                    if not isinstance(cell_l, MergedCell):
                        cell_l.value = label_text
                    cell_l.font = Font(name='Calibri', bold=True, size=7, color="FFFFFF")
                    cell_l.alignment = Alignment(horizontal='center', vertical='center')

            if end_col > start_col:
                try: ws.merge_cells(start_row=task_row, start_column=start_col, end_row=task_row, end_column=end_col)
                except: pass
                try: ws.merge_cells(start_row=label_row, start_column=start_col, end_row=label_row, end_column=end_col)
                except: pass

        # --- CONFIGURACIÓN DE PÁGINA Y ENCABEZADOS (Fidelity Style) ---
        ws.sheet_view.showGridLines = False
        
        # Mapeo dinámico inyectando columnas divisorias reales entre días
        col_map = {}
        divider_cols = []
        
        current_col = 2
        last_date = None
        for h_idx, hour in enumerate(time_columns):
            curr_date = hour.date()
            if last_date is not None and curr_date != last_date:
                # Insertar columna divisoria real para transición entre días
                date_str = f"--- {hour.strftime('%d %b').upper()} ---"
                divider_cols.append({
                    'col': current_col,
                    'date_str': date_str
                })
                current_col += 1
            col_map[hour] = current_col
            current_col += COLS_PER_HOUR
            last_date = curr_date
            
        grid_width = current_col - 2
        
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles import Color

        # --- CABECERA PREMIUM (Rows 1-2) ---
        header_bg = PatternFill("solid", fgColor="F8FAFC")
        for r in range(1, 4):
            for c in range(1, 2 + grid_width):
                ws.cell(row=r, column=c).fill = header_bg

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20
        
        c_title = ws.cell(row=1, column=1)
        font_black = InlineFont(); font_black.rFont = 'Calibri'; font_black.b = True; font_black.sz = 16.0; font_black.color = Color(rgb="0F172A")
        font_blue = InlineFont(); font_blue.rFont = 'Calibri'; font_blue.b = True; font_blue.sz = 16.0; font_blue.color = Color(rgb="2563EB")
        c_title.value = CellRichText([TextBlock(font_black, "Planificación "), TextBlock(font_blue, "Visual")])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
        c_title.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        c_sub = ws.cell(row=2, column=1)
        c_sub.value = "Control de línea ABBAMAT"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)
        c_sub.font = Font(name='Calibri', size=9, color="64748B")
        c_sub.alignment = Alignment(horizontal='left', vertical='top', indent=1)

        # Configuración de cabeceras de día y hora
        header_fill = PatternFill("solid", fgColor="2C3E50")
        header_border = Border(
            left=Side(style='thin', color="475569"), right=Side(style='thin', color="475569"),
            top=Side(style='medium', color="2C3E50"), bottom=Side(style='thin', color="475569")
        )
        ROW_DAY = 3
        ROW_HOUR = 4
        DATA_START = 5

        ws.row_dimensions[ROW_DAY].height = 30
        ws.row_dimensions[ROW_HOUR].height = 20

        # Encabezado MÁQUINA en A3:A4
        c_maq = ws.cell(row=ROW_DAY, column=1)
        if not isinstance(c_maq, MergedCell):
            c_maq.value = "MÁQUINA"
        ws.merge_cells(start_row=ROW_DAY, start_column=1, end_row=ROW_HOUR, end_column=1)
        for r_idx in [ROW_DAY, ROW_HOUR]:
            cell = ws.cell(row=r_idx, column=1)
            cell.fill = header_fill
            cell.border = header_border
        c_maq.font = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
        c_maq.alignment = Alignment(horizontal='center', vertical='center')

        # Calcular rangos de columnas por día
        day_ranges = {}
        for hour, col in col_map.items():
            dt = hour.date()
            if dt not in day_ranges:
                day_ranges[dt] = {'start': col, 'end': col + COLS_PER_HOUR - 1}
            else:
                day_ranges[dt]['end'] = col + COLS_PER_HOUR - 1

        # Dibujar horas
        for hour, h_col in col_map.items():
            c_h = ws.cell(row=ROW_HOUR, column=h_col)
            if not isinstance(c_h, MergedCell):
                c_h.value = hour.strftime("%H")
            ws.merge_cells(start_row=ROW_HOUR, start_column=h_col, end_row=ROW_HOUR, end_column=h_col + COLS_PER_HOUR - 1)
            for offset in range(COLS_PER_HOUR):
                cell = ws.cell(row=ROW_HOUR, column=h_col + offset)
                cell.fill = header_fill
                cell.border = header_border
            c_h.alignment = Alignment(horizontal='center', vertical='center')
            c_h.font = Font(name='Calibri', bold=True, size=9, color="FFFFFF")

        # Dibujar días
        for dt, rng in day_ranges.items():
            start_m = rng['start']
            end_m = rng['end']
            d_str = dt.strftime("%d %b - %a").upper()
            c_d = ws.cell(row=ROW_DAY, column=start_m)
            c_d.value = d_str
            ws.merge_cells(start_row=ROW_DAY, start_column=start_m, end_row=ROW_DAY, end_column=end_m)
            for col_idx in range(start_m, end_m + 1):
                cell = ws.cell(row=ROW_DAY, column=col_idx)
                cell.fill = header_fill
                cell.border = header_border
            c_d.font = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
            c_d.alignment = Alignment(horizontal='center', vertical='center')

        # Dibujar y dar estilo a las cabeceras de columnas divisorias
        for div in divider_cols:
            col_idx = div['col']
            ws.merge_cells(start_row=ROW_DAY, start_column=col_idx, end_row=ROW_HOUR, end_column=col_idx)
            c_div_hdr = ws.cell(row=ROW_DAY, column=col_idx)
            c_div_hdr.value = div['date_str']
            c_div_hdr.fill = PatternFill("solid", fgColor="1E293B")
            c_div_hdr.font = Font(name='Calibri', bold=True, italic=True, size=9, color="FFFFFF")
            c_div_hdr.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
            for r in [ROW_DAY, ROW_HOUR]:
                ws.cell(row=r, column=col_idx).border = Border(
                    left=Side(style='medium', color="1E293B"),
                    right=Side(style='medium', color="1E293B"),
                    top=Side(style='medium', color="1E293B") if r == ROW_DAY else None,
                    bottom=Side(style='medium', color="1E293B") if r == ROW_HOUR else None
                )

        # Auto-ajustar ancho de columna A
        max_len = 15
        for row_data in timeline_data:
            m_name = row_data['machine'].nombre or ''
            if len(m_name) > max_len:
                max_len = len(m_name)
        ws.column_dimensions['A'].width = max_len + 5

        # Establecer anchos de columna de línea de tiempo
        from openpyxl.utils import get_column_letter
        for c in range(2, 2 + grid_width):
            is_div = any(d['col'] == c for d in divider_cols)
            if is_div:
                ws.column_dimensions[get_column_letter(c)].width = 5
            else:
                ws.column_dimensions[get_column_letter(c)].width = 2.5
        
        # --- RENDER DE DATOS (desde DATA_START) ---
        current_row = DATA_START
        divider_fill = PatternFill("solid", fgColor="334155")
        divider_font = Font(name='Calibri', bold=True, italic=True, size=9, color="FFFFFF")
        divider_align = Alignment(horizontal='center', vertical='center', text_rotation=90)
        grid_border = Border(
            left=Side(style='thin', color="D2D2D2"),
            right=Side(style='thin', color="D2D2D2"),
            top=Side(style='thin', color="D2D2D2"),
            bottom=Side(style='thin', color="D2D2D2")
        )

        for idx, row_data in enumerate(timeline_data):
            maquina = row_data['machine']
            tasks = row_data['tasks']
            if maquina.nombre.upper() == 'SIN ASIGNAR' and not tasks: continue
            
            l_row, t_row = current_row, current_row + 1
            ws.row_dimensions[l_row].height = 14
            ws.row_dimensions[t_row].height = 42
            
            c_n = ws.cell(row=l_row, column=1)
            if not isinstance(c_n, MergedCell):
                c_n.value = maquina.nombre.upper()
            ws.merge_cells(start_row=l_row, start_column=1, end_row=t_row, end_column=1)
            c_n.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c_n.font = Font(name='Calibri', bold=True, size=9, color="1E293B")
            c_n.fill = PatternFill("solid", fgColor="F8F9FA")
            c_n.border = Border(bottom=Side(style='thin', color="E2E8F0"), right=Side(style='thin', color="E2E8F0"))
            
            is_odd_row = (idx % 2 == 1)
            zebra_fill = PatternFill("solid", fgColor="F8FAFC") if is_odd_row else PatternFill("solid", fgColor="FFFFFF")
            
            # Llenar la grilla vacía e inyectar divisiones físicas de días
            for c in range(2, 2 + grid_width):
                is_div = any(d['col'] == c for d in divider_cols)
                if is_div:
                    div_data = next(d for d in divider_cols if d['col'] == c)
                    try:
                        ws.merge_cells(start_row=l_row, start_column=c, end_row=t_row, end_column=c)
                    except:
                        pass
                    cell = ws.cell(row=l_row, column=c)
                    cell.value = div_data['date_str']
                    cell.fill = divider_fill
                    cell.font = divider_font
                    cell.alignment = divider_align
                    for r in [l_row, t_row]:
                        ws.cell(row=r, column=c).border = Border(
                            left=Side(style='medium', color="1E293B"),
                            right=Side(style='medium', color="1E293B"),
                            top=Side(style='thin', color="334155") if r == l_row else None,
                            bottom=Side(style='thin', color="334155") if r == t_row else None
                        )
                else:
                    cell_l = ws.cell(row=l_row, column=c)
                    cell_l.fill = zebra_fill
                    cell_l.border = grid_border
                    cell_t = ws.cell(row=t_row, column=c)
                    cell_t.fill = zebra_fill
                    cell_t.border = grid_border
            
            for t in tasks:
                if not isinstance(t, dict):
                    continue
                start_date = t.get('start_date')
                if not start_date: continue
                slot_hour = next((h for h in time_columns if h.date() == start_date.date() and h.hour == start_date.hour), None)
                if not slot_hour: continue
                
                h_col = col_map[slot_hour]
                m_off = int(start_date.minute / 10.0)
                s_col = h_col + m_off
                e_col = s_col + int(t.get('duration_real', 0) * COLS_PER_HOUR)
                
                if s_col < 2: s_col = 2
                if e_col > 2 + grid_width: e_col = 2 + grid_width
                
                if s_col < e_col:
                    draw_floating_card(ws, l_row, t_row, s_col, e_col - 1, 
                                       proyecto_color_map.get(t.get('ProyectoCode'), '0078D4'), 
                                       t.get('ProyectoCode', 'S/P'), t.get('Idorden', ''),
                                       is_critical=t.get('is_critical', False), 
                                       is_delayed=t.get('is_delayed', False), 
                                       delay_days=t.get('delay_days', 0),
                                       is_continuation=t.get('segment_index', 0) > 0)
            current_row += 2

        ws.freeze_panes = f'B{DATA_START}'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Gantt_Produccion.xlsx'
        wb.save(response)
        return response

    except Exception as global_err:
        import traceback
        print(traceback.format_exc())
        return HttpResponse(f"Error critico en Exportacion Excel: {str(global_err)}", status=500)



@csrf_exempt
def create_scenario(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        import json
        
        # INYECCIÓN DE DEBUGGING: Rastrear payload crudo recibido
        print("====== DEBUGGING GUARDAR ORDEN MANUAL ======")
        print("PAYLOAD RECIBIDO RAW:", request.body.decode('utf-8'))
        
        data = json.loads(request.body)
        print("PAYLOAD JSON PARSEADO:", data)
        print("============================================")
        
        nombre = data.get('nombre')
        descripcion = data.get('descripcion', '')
        es_principal = data.get('es_principal', False)
        proyectos = data.get('proyectos', '')
        fecha_inicio_str = data.get('fecha_inicio')
        copy_from_id = data.get('copy_from_id')
        scenario_id = data.get('id') or data.get('update_id')
        secuencias = data.get('secuencias', [])
        plan_mode_payload = str(data.get('plan_mode') or request.session.get('last_plan_mode', 'manual')).lower()
        persistir_prioridad_manual = plan_mode_payload != 'manual'
        
        print("\n" + "="*80)
        print("[DEBUG ENTRADA] Payload recibido en /api/scenarios/create/:")
        print(f"  Total secuencias recibidas: {len(secuencias)}")
        if secuencias:
            print(f"  Muestra primera secuencia (Frontend): {secuencias[0]}")
            print(f"  Muestra última secuencia (Frontend): {secuencias[-1]}")
        print("="*80 + "\n")
        
        with transaction.atomic(using='default'):
            if es_principal:
                Scenario.objects.using('default').filter(es_principal=True).update(es_principal=False)
                
            if scenario_id:
                # Update existing
                scenario = Scenario.objects.using('default').get(pk=scenario_id)
                if nombre:
                    scenario.nombre = nombre
                if descripcion:
                    scenario.descripcion = descripcion
                scenario.es_principal = es_principal
                scenario.proyectos = proyectos
                if fecha_inicio_str:
                    try:
                        scenario.fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                scenario.save(using='default')
                
                # If we are "overwriting" (copying data from another scenario)
                if copy_from_id and str(copy_from_id) != str(scenario_id):
                    # Clean target scenario first
                    PrioridadManual.objects.using('default').filter(scenario=scenario).delete()
                    ProyectoPrioridad.objects.using('default').filter(scenario=scenario).delete()
                    HiddenTask.objects.using('default').filter(scenario=scenario).delete()
                    PlannedTask.objects.using('default').filter(scenario=scenario).delete()
                    
                    # Clone from source
                    source = Scenario.objects.using('default').get(pk=copy_from_id)
                    
                    # Clone Overrides
                    overrides = PrioridadManual.objects.using('default').filter(scenario=source)
                    new_overrides = [
                        PrioridadManual(
                            id_orden=o.id_orden, maquina=o.maquina, prioridad=o.prioridad,
                            tiempo_manual=o.tiempo_manual, nivel_manual=o.nivel_manual,
                            porcentaje_solapamiento=o.porcentaje_solapamiento,
                            fecha_inicio_manual=o.fecha_inicio_manual,
                            orden_secuencia=o.orden_secuencia,
                            scenario=scenario
                        ) for o in overrides
                    ]
                    PrioridadManual.objects.using('default').bulk_create(new_overrides)

                    # Clone Hidden Tasks
                    hidden = HiddenTask.objects.using('default').filter(scenario=source)
                    new_hidden = [
                        HiddenTask(id_orden=h.id_orden, scenario=scenario)
                        for h in hidden
                    ]
                    HiddenTask.objects.using('default').bulk_create(new_hidden)
                    
                    # Clone Project Priorities
                    proj_prios = ProyectoPrioridad.objects.using('default').filter(scenario=source)
                    new_proj_prios = [
                        ProyectoPrioridad(proyecto=p.proyecto, prioridad=p.prioridad, scenario=scenario)
                        for p in proj_prios
                    ]
                    ProyectoPrioridad.objects.using('default').bulk_create(new_proj_prios)

                    # Clone Planned Tasks
                    planned = PlannedTask.objects.using('default').filter(scenario=source)
                    new_planned = [
                        PlannedTask(id_orden=p.id_orden, scenario=scenario, proyecto_code=p.proyecto_code)
                        for p in planned
                    ]
                    PlannedTask.objects.using('default').bulk_create(new_planned)

                if secuencias:
                    # RECOVERY/MAPPING LOGIC: Ensure all OPs have a PlannedTask with their correct project_code
                    id_ordens = [seq.get('id_orden') for seq in secuencias if seq.get('id_orden')]
                    existing_planned = set(PlannedTask.objects.using('default').filter(
                        scenario=scenario, id_orden__in=id_ordens
                    ).values_list('id_orden', flat=True))
                    
                    missing_ids = [int(oid) for oid in id_ordens if int(oid) not in existing_planned]
                    if missing_ids:
                        from django.db import connections
                        with connections['production'].cursor() as cursor:
                            placeholders = ', '.join(['%s'] * len(missing_ids))
                            cursor.execute(f"SELECT Idorden, Formula FROM Tman050 WHERE Idorden IN ({placeholders})", missing_ids)
                            missing_projects = {row[0]: str(row[1]).strip() if row[1] else '' for row in cursor.fetchall()}
                            
                        new_tasks = []
                        for oid in missing_ids:
                            proj_code = missing_projects.get(oid, '')
                            new_tasks.append(PlannedTask(id_orden=oid, scenario=scenario, proyecto_code=proj_code))
                        PlannedTask.objects.using('default').bulk_create(new_tasks)

                    for seq in secuencias:
                        id_orden = seq.get('id_orden')
                        maquina = seq.get('maquina')
                        # Extracción segura de nivel de planificación
                        orden_secuencia = seq.get('orden_secuencia', 0)
                        raw_nivel = None
                        if persistir_prioridad_manual:
                            raw_nivel = seq.get('nivel_planificacion')
                            if raw_nivel is None:
                                raw_nivel = seq.get('prioridad_manual')
                            
                        # Si encontramos un valor para el nivel en el POST, lo parseamos
                        nivel_final = None
                        if raw_nivel is not None and str(raw_nivel).strip() != '':
                            try:
                                nivel_final = int(raw_nivel)
                            except ValueError:
                                pass
                        
                        # Extraer cantidad_producida y tiempo_proceso del payload
                        cantidad_producida_manual = seq.get('cantidad_producida_manual')
                        tiempo_manual = seq.get('tiempo_manual')
                        
                        print(f"[GUARDANDO] OP: {id_orden} | Maquina: {maquina} | plan_mode: {plan_mode_payload} | raw_nivel: {repr(raw_nivel)} | nivel_final: {nivel_final} | seq_keys: {list(seq.keys())}")
                        
                        if id_orden and maquina:
                            # Calculamos la prioridad multiplicando por 1000 al igual que en api_guardar_orden_manual
                            prioridad_val = (orden_secuencia + 1) * 1000.0
                            
                            defaults_dict = {
                                'orden_secuencia': orden_secuencia,
                                'prioridad': prioridad_val
                            }
                            # Solo pisamos el nivel_manual y la prioridad en la BD si vino un valor válido en el POST
                            if nivel_final is not None:
                                defaults_dict['nivel_manual'] = nivel_final
                                # SE AGREGA: Actualizar también 'prioridad' basándose en el payload de guardado manual
                                defaults_dict['prioridad'] = float(str(nivel_final).replace(',', '.'))
                            if cantidad_producida_manual is not None:
                                defaults_dict['cantidad_producida_manual'] = cantidad_producida_manual
                            if tiempo_manual is not None:
                                defaults_dict['tiempo_manual'] = tiempo_manual
                            
                            print(f"[DEBUG BD] Escribiendo en BD para OP {id_orden}: {defaults_dict}")
                                
                            PrioridadManual.objects.using('default').update_or_create(
                                scenario=scenario,
                                id_orden=id_orden,
                                maquina=maquina,
                                defaults=defaults_dict
                            )

                return JsonResponse({'status': 'ok', 'scenario': {'id': scenario.id, 'nombre': scenario.nombre}})
                
            else:
                # Create NEW scenario (con unicidad por nombre para evitar duplicados)
                if not nombre:
                    return JsonResponse({'error': 'Nombre es requerido'}, status=400)

                fecha_inicio_obj = None
                if fecha_inicio_str:
                    try:
                        fecha_inicio_obj = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                    except ValueError:
                        pass

                # FIX: Si ya existe un escenario con ese nombre, lo reutilizamos en lugar de crear duplicados
                new_scenario, _created = Scenario.objects.using('default').get_or_create(
                    nombre=nombre,
                    defaults={
                        'descripcion': descripcion,
                        'es_principal': es_principal,
                        'proyectos': proyectos,
                        'fecha_inicio': fecha_inicio_obj,
                    }
                )
                
                # If we retrieved an existing scenario rather than creating it, update the fecha_inicio if provided
                if not _created and fecha_inicio_obj:
                    new_scenario.fecha_inicio = fecha_inicio_obj
                    new_scenario.save(using='default')
                
                # Clone overrides if requested
                if copy_from_id:
                    source = Scenario.objects.using('default').get(pk=copy_from_id)
                    overrides = PrioridadManual.objects.using('default').filter(scenario=source)
                    new_overrides = []
                    for o in overrides:
                        new_overrides.append(PrioridadManual(
                            id_orden=o.id_orden, maquina=o.maquina, prioridad=o.prioridad,
                            tiempo_manual=o.tiempo_manual, nivel_manual=o.nivel_manual,
                            porcentaje_solapamiento=o.porcentaje_solapamiento,
                            fecha_inicio_manual=o.fecha_inicio_manual, 
                            orden_secuencia=o.orden_secuencia,
                            scenario=new_scenario
                        ))
                    PrioridadManual.objects.using('default').bulk_create(new_overrides)
                    
                    # Clone Prioridades de Proyecto
                    p_priorities = ProyectoPrioridad.objects.using('default').filter(scenario=source)
                    new_p_priorities = []
                    for p in p_priorities:
                        new_p_priorities.append(ProyectoPrioridad(
                            scenario=new_scenario, proyecto=p.proyecto, prioridad=p.prioridad
                        ))
                    ProyectoPrioridad.objects.using('default').bulk_create(new_p_priorities)

                    # Clone Planned Tasks
                    planned = PlannedTask.objects.using('default').filter(scenario=source)
                    new_planned = [
                        PlannedTask(id_orden=p.id_orden, scenario=new_scenario, proyecto_code=p.proyecto_code)
                        for p in planned
                    ]
                    PlannedTask.objects.using('default').bulk_create(new_planned)
                
                if secuencias:
                    # RECOVERY/MAPPING LOGIC: Ensure all OPs have a PlannedTask with their correct project_code
                    id_ordens = [seq.get('id_orden') for seq in secuencias if seq.get('id_orden')]
                    existing_planned = set(PlannedTask.objects.using('default').filter(
                        scenario=new_scenario, id_orden__in=id_ordens
                    ).values_list('id_orden', flat=True))
                    
                    missing_ids = [int(oid) for oid in id_ordens if int(oid) not in existing_planned]
                    if missing_ids:
                        from django.db import connections
                        with connections['production'].cursor() as cursor:
                            placeholders = ', '.join(['%s'] * len(missing_ids))
                            cursor.execute(f"SELECT Idorden, Formula FROM Tman050 WHERE Idorden IN ({placeholders})", missing_ids)
                            missing_projects = {row[0]: str(row[1]).strip() if row[1] else '' for row in cursor.fetchall()}
                            
                        new_tasks = []
                        for oid in missing_ids:
                            proj_code = missing_projects.get(oid, '')
                            new_tasks.append(PlannedTask(id_orden=oid, scenario=new_scenario, proyecto_code=proj_code))
                        PlannedTask.objects.using('default').bulk_create(new_tasks)

                    for seq in secuencias:
                        id_orden = seq.get('id_orden')
                        maquina = seq.get('maquina')
                        # Extracción segura de nivel de planificación
                        orden_secuencia = seq.get('orden_secuencia', 0)
                        raw_nivel = None
                        if persistir_prioridad_manual:
                            raw_nivel = seq.get('nivel_planificacion')
                            if raw_nivel is None:
                                raw_nivel = seq.get('prioridad_manual')
                            
                        # Si encontramos un valor para el nivel en el POST, lo parseamos
                        nivel_final = None
                        if raw_nivel is not None and str(raw_nivel).strip() != '':
                            try:
                                nivel_final = int(raw_nivel)
                            except ValueError:
                                pass
                        
                        # Extraer cantidad_producida y tiempo_proceso del payload
                        cantidad_producida_manual = seq.get('cantidad_producida_manual')
                        tiempo_manual = seq.get('tiempo_manual')
                        
                        if id_orden and maquina:
                            defaults_dict = {
                                'orden_secuencia': orden_secuencia
                            }
                            # Solo pisamos el nivel_manual y la prioridad en la BD si vino un valor válido en el POST
                            if nivel_final is not None:
                                defaults_dict['nivel_manual'] = nivel_final
                                # SE AGREGA: Actualizar también 'prioridad' basándose en el payload de guardado manual
                                defaults_dict['prioridad'] = float(str(nivel_final).replace(',', '.'))
                            if cantidad_producida_manual is not None:
                                defaults_dict['cantidad_producida_manual'] = cantidad_producida_manual
                            if tiempo_manual is not None:
                                defaults_dict['tiempo_manual'] = tiempo_manual
                                
                            PrioridadManual.objects.using('default').update_or_create(
                                scenario=new_scenario,
                                id_orden=id_orden,
                                maquina=maquina,
                                defaults=defaults_dict
                            )

                return JsonResponse({
                    'status': 'ok',
                    'scenario': {'id': new_scenario.id, 'nombre': new_scenario.nombre}
                })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def delete_scenario(request, scenario_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        scenario = get_object_or_404(Scenario, pk=scenario_id)
        
        if scenario.es_principal:
            return JsonResponse({'error': 'No se puede eliminar el Plan Oficial'}, status=400)
            
        scenario.delete()
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def api_list_scenarios(request):
    """
    Returns a list of all scenarios in the database.
    """
    try:
        from .models import Scenario
        scenarios = Scenario.objects.using('default').all().order_by('-id')
        data = []
        for s in scenarios:
            data.append({
                'id': s.id,
                'nombre': s.nombre,
                'es_principal': s.es_principal,
                'proyectos': s.proyectos or ''
            })
        return JsonResponse({'scenarios': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def publish_scenario(request, scenario_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        with transaction.atomic():
            # Set all to False
            Scenario.objects.using('default').update(es_principal=False)
            
            # Set target to True
            target = Scenario.objects.using('default').get(pk=scenario_id)
            target.es_principal = True
            target.save(using='default')
            
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def estadisticas_produccion(request):
    """
    Dashboard for system-wide statistics and machine occupancy.
    """
    from .gantt_logic import get_gantt_data
    # Force run of shared logic to get statistics
    # Note: Using default behavior (next 7 days lookahead)
    data = get_gantt_data(request, force_run=True)
    
    analysis = data.get('analysis', {})
    machines = analysis.get('machines', [])
    
    # Filter out machines with 0 capacity to avoid division by zero
    total_capacity = sum(m['capacity'] for m in machines)
    total_hours = sum(m['hours'] for m in machines)
    avg_load = (total_hours / total_capacity * 100) if total_capacity > 0 else 0
    
    # Metrics
    collapsed_machines = [m for m in machines if m['load_pct'] >= 100]
    high_load_machines = [m for m in machines if 70 <= m['load_pct'] < 100]
    healthy_machines = [m for m in machines if m['load_pct'] < 50]
    
    # Recent Alerts
    project_alerts = analysis.get('project_alerts', [])
    
    context = {
        'analysis': analysis,
        'machines': machines,
        'total_capacity': round(total_capacity, 1),
        'total_hours': round(total_hours, 1),
        'avg_load': round(avg_load, 1),
        'collapsed_count': len(collapsed_machines),
        'high_load_count': len(high_load_machines),
        'healthy_count': len(healthy_machines),
        'collapsed_machines': collapsed_machines,
        'healthy_machines': healthy_machines,
        'project_alerts': project_alerts,
        'active_scenario': data.get('active_scenario'),
        'today': datetime.now(),
    }
    
    return render(request, 'produccion/estadisticas.html', context)


def proyectos_prioridades(request):
    scenario_id = request.GET.get('scenario_id')
    active_scenario = get_active_scenario(request, scenario_id=scenario_id)

    # MISMA FUENTE que la grilla principal: T2.Formula del ERP
    # La lista lateral derecha usa ProyectoCode = T2.Formula para mostrar
    # los proyectos activos. Reproducimos exactamente esa lógica aquí.
    proyectos_list = []
    if active_scenario:
        planned_ids = list(PlannedTask.objects.using('default').filter(
            scenario=active_scenario
        ).values_list('id_orden', flat=True))

        if planned_ids:
            from django.db import connections
            placeholders = ', '.join(['%s'] * len(planned_ids))
            with connections['production'].cursor() as cursor:
                cursor.execute(f"""
                    SELECT DISTINCT T2.Formula
                    FROM Tman050 T
                    INNER JOIN Tman050 T2 ON T.MSTNMBR = T2.IdOrden
                    WHERE T.Idorden IN ({placeholders})
                      AND T2.Formula IS NOT NULL
                      AND LTRIM(RTRIM(T2.Formula)) != ''
                    ORDER BY T2.Formula
                """, planned_ids)
                proyectos_list = [str(row[0]).strip() for row in cursor.fetchall() if row[0]]

    # AUTO-VERIFICACIÓN: comparar con PlannedTask.proyecto_code directo
    planned_direct = []
    if active_scenario:
        planned_direct = list(PlannedTask.objects.using('default').filter(
            scenario=active_scenario
        ).values_list('proyecto_code', flat=True).distinct().order_by('proyecto_code'))
    if set(proyectos_list) != set(planned_direct):
        solo_erp = set(proyectos_list) - set(planned_direct)
        solo_sqlite = set(planned_direct) - set(proyectos_list)
        print(f"DEBUG PRIORIDADES: DIFERENCIA entre ERP (T2.Formula) y SQLite (proyecto_code)!")
        if solo_erp:
            print(f"  Solo en ERP: {sorted(solo_erp)}")
        if solo_sqlite:
            print(f"  Solo en SQLite: {sorted(solo_sqlite)}")

    print(f"DEBUG PRIORIDADES: {len(proyectos_list)} proyectos desde ERP (T2.Formula): {proyectos_list}")

    if not proyectos_list:
        return render(request, 'produccion/proyectos_prioridades.html', {
            'proyectos_data': [],
            'active_scenario': active_scenario,
            'all_scenarios': Scenario.objects.using('default').all().order_by('-fecha_creacion'),
            'has_projects': False
        })

    prioridades_db = {}
    if active_scenario:
        db_qs = ProyectoPrioridad.objects.using('default').filter(
            scenario=active_scenario,
            proyecto__in=proyectos_list
        )
        prioridades_db = {p.proyecto: p.prioridad for p in db_qs}

    proyectos_data = []
    for p in proyectos_list:
        prio = prioridades_db.get(p, 999)
        proyectos_data.append({'proyecto': p, 'prioridad': prio})

    assigned_prios = [d['prioridad'] for d in proyectos_data if d['prioridad'] != 999]
    max_prio = max(assigned_prios + [0])
    for item in proyectos_data:
        if item['prioridad'] == 999:
            max_prio += 1
            item['prioridad'] = max_prio

    proyectos_final = proyectos_data
    print(f"DEBUG PRIORIDADES: enviando {len(proyectos_final)} proyectos a la plantilla")
    print(f"DEBUG PRIORIDADES: ESPEJO SIDEBAR OK? {set(proyectos_list) == set(planned_direct)}")

    return render(request, 'produccion/proyectos_prioridades.html', {
        'proyectos': proyectos_final,
        'active_scenario': active_scenario,
        'all_scenarios': Scenario.objects.using('default').all().order_by('-fecha_creacion'),
        'has_projects': len(proyectos_final) > 0,
        'total_projects': len(proyectos_final),
    })

@csrf_exempt
def update_proyecto_prioridad(request):
    """
    API endpoint to update the priority of multiple projects for the active scenario.
    Expects a JSON body with a list of updates:
    {
        "scenario_id": 1,
        "updates": [
            {"proyecto": "25-001", "prioridad": 1},
            {"proyecto": "23-145", "prioridad": 2}
        ]
    }
    """
    if request.method != 'POST':
         return JsonResponse({'error': 'Method not allowed'}, status=405)
         
    try:
        data = json.loads(request.body)
        scenario_id = data.get('scenario_id')
        updates = data.get('updates', [])
        
        scenario = None
        if scenario_id:
             scenario = Scenario.objects.using('default').filter(pk=scenario_id).first()
        else:
             scenario = Scenario.objects.using('default').filter(es_principal=True).first()
             
        if not scenario:
             return JsonResponse({'error': 'No active scenario found'}, status=400)
             
        with transaction.atomic(using='default'):
             for update in updates:
                  proyecto = update.get('proyecto')
                  prioridad = update.get('prioridad')
                  
                  if proyecto and prioridad is not None:
                       print(f"POST PRIORIDADES: Procesando Proyecto {proyecto} con nueva prioridad {prioridad}")
                       # Update or create priority
                       ProyectoPrioridad.objects.using('default').update_or_create(
                            scenario=scenario,
                            proyecto=proyecto,
                            defaults={'prioridad': int(prioridad)}
                       )
                       
        return JsonResponse({'status': 'ok'})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


def planillas_diarias(request):
    """
    Generates a daily production sheet per machine based on Gantt calculation.
    """
    from .gantt_logic import get_gantt_data
    from collections import defaultdict
    from datetime import datetime, timedelta
    
    # 1. OBTENCIÓN DE DATOS LIMPIA
    gantt_res = get_gantt_data(request, force_run=True)
    timeline_data = gantt_res.get('timeline_data', [])
    
    daily_plan = {}
    active_dates = set()
    
    for machine_row in timeline_data:
        machine = machine_row['machine']
        m_id = str(machine.id_maquina)
        
        if m_id == 'MAC00' and not machine_row['tasks']:
            continue
            
        daily_plan[m_id] = {
            'machine_id': m_id,
            'machine_name': machine.nombre,
            'dates': defaultdict(list)
        }
        
        task_segments = []
        for segment in machine_row['tasks']:
            start = segment.get('start_date')
            if not start: continue
            task_segments.append(segment)
            
        segments_per_task = defaultdict(int)
        for seg in task_segments:
            tid = str(seg.get('Idorden'))
            segments_per_task[tid] += 1
            
        current_segment_index = defaultdict(int)
        processed_qty_map = defaultdict(float) 
        
        for segment in task_segments:
            start = segment.get('start_date')
            if start.weekday() >= 5: 
                continue
                
            d_str = start.date().isoformat()
            t_id = str(segment.get('Idorden'))
            
            cant_total = float(segment.get('cantidad_final') or segment.get('Cantidad_Final') or 0.0)
            cant_prod = float(segment.get('cantidad_producida') or segment.get('Cantidadpp') or 0.0)
            
            qty_pend = cant_total - cant_prod
            if qty_pend < 0: qty_pend = 0.0
            
            total_duration = float(segment.get('Tiempo_Proceso') or 0.001)
            if total_duration <= 0: total_duration = 0.001
            
            segment_duration = float(segment.get('duration_real') or 0.0)
            
            current_segment_index[t_id] += 1
            is_last_segment = current_segment_index[t_id] == segments_per_task[t_id]
            
            if is_last_segment:
                segment_qty = qty_pend - processed_qty_map[t_id]
            else:
                raw_qty = qty_pend * (segment_duration / total_duration)
                segment_qty = round(raw_qty * 2.0) / 2.0
                if processed_qty_map[t_id] + segment_qty > qty_pend:
                    segment_qty = qty_pend - processed_qty_map[t_id]
            
            if segment_qty < 0: segment_qty = 0.0
            processed_qty_map[t_id] += segment_qty
            
            std_t = float(segment.get('Tiempo') or 0.0)
            total_std_time = std_t * segment_qty

            if segment_qty > 0 or segment_duration > 0:
                h_tot = int(total_std_time)
                m_tot = int(round((total_std_time - h_tot) * 60))
                if m_tot >= 60:
                    h_tot += 1
                    m_tot = 0
                tiempo_dia_hm = f"{h_tot}:{m_tot:02d}h"

                h_standard = int(std_t)
                m_standard = int(round((std_t - h_standard) * 60))
                if m_standard >= 60:
                    h_standard += 1
                    m_standard = 0
                tiempo_standard_hm = f"{h_standard}:{m_standard:02d}h"

                daily_plan[m_id]['dates'][d_str].append({
                    'orden': t_id,
                    'proyecto': segment.get('ProyectoCode'),
                    'denominacion': segment.get('Denominacion'),
                    'descripcion': segment.get('Descri'),
                    'tiempo_standard': tiempo_standard_hm, 
                    'cantidad_dia': segment_qty,
                    'tiempo_dia': tiempo_dia_hm,
                    'start_time': start.strftime('%H:%M'),
                    'end_time': segment.get('end_date').strftime('%H:%M') if segment.get('end_date') else ''
                })
                active_dates.add(d_str)
                
    for m_id, m_data in daily_plan.items():
        m_data['dates'] = dict(m_data['dates'])
        for d in m_data['dates']:
            m_data['dates'][d].sort(key=lambda x: x['start_time'])
        m_data['has_active_dates'] = len(m_data['dates']) > 0
        
    daily_plan_list = [v for k, v in daily_plan.items() if v.get('has_active_dates')]

    # 2. ENCABEZADOS DINÁMICOS CON CORTE DE SEMANA ESTRICTO
    DAYS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    nice_target_dates = []
    
    try:
        semanas = int(request.GET.get('semanas', 1))
    except ValueError:
        semanas = 1
        
    active_scenario = gantt_res.get('active_scenario')
    if active_scenario and active_scenario.fecha_inicio:
        base_date = active_scenario.fecha_inicio
    else:
        # Fallback to earliest active date or today
        if active_dates:
            sorted_dates = sorted([datetime.fromisoformat(d) for d in active_dates])
            base_date = sorted_dates[0].date()
        else:
            base_date = datetime.now().date()
            
    curr = base_date
    end_date = base_date + timedelta(days=(semanas * 7) - 1)
    
    while curr <= end_date:
        if curr.weekday() < 5: 
            nice_target_dates.append((curr.isoformat(), f"{DAYS_ES[curr.weekday()]} {curr.strftime('%d/%m')}"))
        curr += timedelta(days=1)
    
    context = {
        'daily_plan': daily_plan_list,
        'target_dates': nice_target_dates,
        'active_scenario': gantt_res.get('active_scenario'),
        'semanas': semanas
    }
    
    return render(request, 'produccion/planillas_diarias.html', context)

@csrf_exempt
def redistribute_tasks(request):
    """
    API endpoint para redistribuir SOLO las tareas que se solapan con la falla
    de la máquina hacia otra máquina compatible.
    """
    from django.http import JsonResponse
    from .models import PrioridadManual, Scenario, MaquinaEquivalencia, HiddenTask
    from .services import get_planificacion_data
    from django.utils import timezone
    from django.db.models import Max

    try:
        from_machine_id = request.GET.get('from')
        to_machine_id = request.GET.get('to')
        proyectos_p = request.GET.get('proyectos')
        scenario_id = request.GET.get('scenario_id')

        if not from_machine_id or not to_machine_id:
            return JsonResponse({'success': False, 'error': 'Parámetros incompletos'}, status=400)

        # 1. Resolver Escenario
        scenario = Scenario.objects.using('default').filter(pk=scenario_id).first() if scenario_id else None
        if not scenario:
            scenario = Scenario.objects.using('default').filter(es_principal=True).first()
            
        # NUEVA LÓGICA: Validar si hay tareas planificadas en lugar de exigir el objeto escenario
        if not scenario:
            from .models import PlannedTask
            has_tasks = PlannedTask.objects.using('default').filter(scenario__isnull=True).exists()
            if not has_tasks:
                return JsonResponse({'success': False, 'error': 'No hay tareas en la planificación actual'}, status=400)

        # 2. Buscar Factor de Eficiencia por Equivalencia
        equivalencia = MaquinaEquivalencia.objects.using('default').filter(
            maquina_origen_id=from_machine_id,
            maquina_destino_id=to_machine_id
        ).first()
        factor = equivalencia.factor_eficiencia if equivalencia else 1.0

        # 3. Obtener Tareas del Origen (usando el Gantt actual para ver qué está realmente allí)
        print(f"DEBUG: [Redistribute] From: {from_machine_id}, To: {to_machine_id}, Proyectos: {proyectos_p}")
        
        from .gantt_logic import get_gantt_data
        from datetime import timedelta
        
        class MockRequest:
            def __init__(self, projects, s_id):
                self.GET = {
                    'run': '1',
                    'proyectos': projects or '',
                    'scenario_id': str(s_id) if s_id else '',
                    'plan_mode': 'manual'
                }
                self.session = {}
        
        mock_req = MockRequest(proyectos_p, scenario.id)
        gantt_data = get_gantt_data(mock_req, force_run=True)
        
        # Obtener rango de la falla para filtrar (Solo Inicio)
        from .models import MantenimientoMaquina, MaquinaConfig
        now = timezone.now()
        failure = MantenimientoMaquina.objects.using('default').filter(
            maquina_id=from_machine_id,
            estado__in=['FALLA', 'EN_CURSO', 'PROGRAMADO'],
            fecha_fin__gte=now
        ).order_by('fecha_inicio').first()
        
        # Si no hay falla activa, el filtro de inicio es hoy (Redistribución por carga)
        f_start = failure.fecha_inicio if failure else now
        
        if failure:
            print(f"DEBUG: [Redistribute] Queue Start: {f_start}")

        affected_tasks = []
        from_m_upper = str(from_machine_id).strip().upper()
        
        all_machine_tasks = []

        for row in gantt_data.get('timeline_data', []):
            m_id = str(row['machine'].id_maquina).strip().upper()
            if m_id == from_m_upper:
                for t in row.get('tasks', []):
                    all_machine_tasks.append(str(t.get('Idorden')))
                    
                    task_start = t.get('start_date')
                    if task_start:
                        # Ensure comparison in same timezone reference
                        if timezone.is_naive(task_start): task_start = timezone.make_aware(task_start)
                        
                        # NUEVA LOGICA: Cola de Producción
                        # Cualquier tarea que empiece DESPUÉS de que inicie la falla (Bloqueo Total)
                        if task_start >= f_start:
                            affected_tasks.append(t)
        
        print(f"DEBUG: [Redistribute] Total Tasks on {from_machine_id}: {len(all_machine_tasks)}")
        print(f"DEBUG: [Redistribute] Affected in Queue: {len(affected_tasks)}")

        if not affected_tasks:
            return JsonResponse({
                'success': False, 
                'error': 'No se encontraron tareas afectadas por la falla en este horario.'
            }, status=200) # Use 200 so UI can show the message instead of alert box crash

        # 4. Agrupar por Proyecto para mantener cohesión
        # Ordenamos primero por ProyectoCode para el agrupamiento
        affected_tasks.sort(key=lambda x: (x.get('ProyectoCode', 'ZZZ'), x.get('OrdenVisual', 0)))

        # 5. Calcular Punto de Inserción (Max Prioridad + 100)
        max_prio = PrioridadManual.objects.using('default').filter(
            scenario=scenario, maquina=to_machine_id
        ).aggregate(Max('prioridad'))['prioridad__max'] or 1000000.0
        
        next_prio = max_prio + 100.0

        # 6. Ejecutar Movimiento
        moved_count = 0
        from django.db import transaction
        with transaction.atomic(using='default'):
            for task in affected_tasks:
                task_id = task.get('Idorden')
                # Tiempo original de pieza (Tiempo_Proceso / Cantidad Pendiente)
                # O mejor: El sistema ya calculó el `Tiempo_Proceso` en el Gantt original
                original_total_time = float(task.get('Tiempo_Proceso', 0) or 0)
                if original_total_time <= 0: continue
                
                # RE-CALCULAR TIEMPO POR EFICIENCIA
                new_total_time = original_total_time * factor
                
                # Borrar asignación previa en este escenario
                PrioridadManual.objects.using('default').filter(id_orden=task_id, scenario=scenario).delete()
                
                # Crear nueva asignación
                PrioridadManual.objects.using('default').create(
                    id_orden=task_id,
                    scenario=scenario,
                    maquina=to_machine_id,
                    prioridad=next_prio,
                    tiempo_manual=new_total_time, # Sobrescribir tiempo por equivalencia
                    porcentaje_solapamiento=task.get('porcentaje_solapamiento', 0.0),
                    nivel_manual=None # Limpiar niveles antiguos para usar solo prioridad
                )
                
                next_prio += 100.0 # Siguiente paso
                moved_count += 1

        return JsonResponse({
            'success': True,
            'moved_count': moved_count,
            'message': f'Se redistribuyeron {moved_count} tareas de {from_machine_id} a {to_machine_id} aplicando un factor de eficiencia x{factor}.'
        })


    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
def api_get_project_articles(request):
    """
    NIVEL 1: Obtiene la lista de artículos principales (Piezas/Conjuntos) para un proyecto.
    Utiliza el flag IsMacro para identificar únicamente las cabeceras.
    """
    proyecto = request.GET.get('proyecto', '').strip()
    scenario_id = request.GET.get('scenario_id')
    active_scenario = get_active_scenario(request, scenario_id=scenario_id)
    
    if not proyecto:
        return JsonResponse({'error': 'Proyecto no especificado'}, status=400)
    
    # 1. Buscamos los artículos principales en el ERP (BUSQUEDA EXACTA)
    # Nota: Filtramos por Formula para capturar la OP master. IsMacro = 1 son cabeceras.
    # IMPORTANTE: Usamos igualdad exacta (=), NO LIKE, para evitar que "26-038" traiga
    # subproyectos como "26-038-G" o "26-038-J".
    sql_articles = """
    SELECT 
        Articulo,
        Descri as Denominacion,
        SUM(Cantidad) as Solicitado,
        SUM(Cantidadpp) as Finalizado,
        MacroPK,
        MAX(Idorden) as IdOrdenMaster
    FROM Tman050
    WHERE Formula = %s
    AND IsMacro = 1
    GROUP BY Articulo, Descri, MacroPK
    ORDER BY Descri
    """
    
    with connections['production'].cursor() as cursor:
        cursor.execute(sql_articles, [proyecto])
        cols = [col[0] for col in cursor.description]
        articles = [dict(zip(cols, row)) for row in cursor.fetchall()]

        # 2. Buscamos qué OPs de este proyecto YA están en la planificación actual
        # Consultamos SQLite para saber qué está planificado en este escenario
        planned_ids = list(PlannedTask.objects.using('default').filter(
            scenario=active_scenario
        ).values_list('id_orden', flat=True))

        planned_state = {}
        if planned_ids:
            # Consultamos el ERP para saber a qué MacroPK pertenecen estas OPs planificadas
            # Usamos placeholders dinámicos para evitar errores de sintaxis con IN
            placeholders = ', '.join(['%s'] * len(planned_ids))
            sql_mapping = f"""
            SELECT MacroPK, Idorden 
            FROM Tman050 
            WHERE Idorden IN ({placeholders})
            AND Formula = %s
            """
            params = list(planned_ids) + [proyecto]
            cursor.execute(sql_mapping, params)
            mapping_rows = cursor.fetchall()
            
            for m_pk, oid in mapping_rows:
                oid_s = str(oid)
                if m_pk not in planned_state:
                    planned_state[m_pk] = []
                planned_state[m_pk].append(oid_s)

        # 3. Consultamos los niveles manuales guardados en SQLite para este escenario
        # Queremos las prioridades tanto de las OPs planificadas como de las Masters OPs (para artículos con 0 seleccionados)
        master_ids = [art.get('IdOrdenMaster') for art in articles if art.get('IdOrdenMaster')]
        ids_to_query = list(set(planned_ids + master_ids))
        
        if ids_to_query:
            p_manual_db = PrioridadManual.objects.using('default').filter(
                scenario=active_scenario,
                id_orden__in=ids_to_query
            ).values('id_orden', 'nivel_manual')
            
            op_to_nivel = {p['id_orden']: p['nivel_manual'] for p in p_manual_db if p['nivel_manual'] is not None}
            
            # Enriquecemos los artículos con su nivel actual
            for art in articles:
                art['nivel_planificacion'] = 0 # Default si no hay override
                m_pk = art.get('MacroPK')
                master_id = art.get('IdOrdenMaster')
                
                # Primero probamos con el master ID (que es donde guardamos la pieza independientemente de si hay seleccionados)
                if master_id and int(master_id) in op_to_nivel:
                    art['nivel_planificacion'] = op_to_nivel[int(master_id)]
                # Fallback: Si alguna OP hija tiene override
                elif m_pk in planned_state:
                    for oid_s in planned_state[m_pk]:
                        if int(oid_s) in op_to_nivel:
                            art['nivel_planificacion'] = op_to_nivel[int(oid_s)]
                            break

    return JsonResponse({
        'articles': articles,
        'planned_state': planned_state
    })

@csrf_exempt
def api_get_article_processes(request):
    """
    NIVEL 2: Obtiene los procesos (OPs) vinculados a un MacroPK específico.
    """
    macro_pk = request.GET.get('macro_pk', '').strip()
    
    if not macro_pk:
        return JsonResponse({'error': 'MacroPK no especificado'}, status=400)
    
    # Buscamos las operaciones vinculadas al MacroPK. 
    # Filtramos IsMacro = 0 para que no se traiga el artículo padre, solo los procesos.
    # Join con Tman010 para traer el nombre de la máquina.
    # Consultamos procesos vinculados al MacroPK (Principal) 
    # MÁS procesos que compartan la misma "Madre" (MSTNMBR) pero que no tengan MacroPK (Huérfanos/Fallback)
    sql = """
    SELECT 
        T.Idorden as IdOrden,
        T.Articulo as Articulo,
        T.Formula as Denominacion,
        T.Descri as Proceso,
        (T.Cantidad - T.Cantidadpp) as Pendiente,
        T.Cantidad as Cantidad,
        T.Cantidadpp as Finalizado,
        ISNULL((SELECT MAX(T3.Nivel) FROM TMAN002 T3 WHERE LTRIM(RTRIM(T3.ArticuloH)) = LTRIM(RTRIM(T.Articulo)) AND LTRIM(RTRIM(T3.Formula)) = LTRIM(RTRIM(T.Formula))), 0) as Nivel_Planificacion,
        ISNULL(M.MAQUINAD, T.Idmaquina) as MaquinaNombre,
        T.MSTNMBR as MSTNMBR
    FROM Tman050 T
    LEFT JOIN Tman010 M ON T.Idmaquina = M.Idmaquina
    WHERE T.MacroPK = %s
    AND T.IsMacro = 0

    UNION

    SELECT 
        T.Idorden as IdOrden,
        T.Articulo as Articulo,
        T.Formula as Denominacion,
        T.Descri as Proceso,
        (T.Cantidad - T.Cantidadpp) as Pendiente,
        T.Cantidad as Cantidad,
        T.Cantidadpp as Finalizado,
        ISNULL((SELECT MAX(T3.Nivel) FROM TMAN002 T3 WHERE LTRIM(RTRIM(T3.ArticuloH)) = LTRIM(RTRIM(T.Articulo)) AND LTRIM(RTRIM(T3.Formula)) = LTRIM(RTRIM(T.Formula))), 0) as Nivel_Planificacion,
        ISNULL(M.MAQUINAD, T.Idmaquina) as MaquinaNombre,
        T.MSTNMBR as MSTNMBR
    FROM Tman050 T
    LEFT JOIN Tman010 M ON T.Idmaquina = M.Idmaquina
    WHERE T.MSTNMBR IN (SELECT IdOrden FROM Tman050 WHERE MacroPK = %s AND IsMacro = 1)
    AND (T.MacroPK IS NULL OR T.MacroPK = '')
    AND T.IsMacro = 0
    
    ORDER BY IdOrden
    """
    
    try:
        with connections['production'].cursor() as cursor:
            cursor.execute(sql, [macro_pk, macro_pk])
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
        # 3. Aplicar Overrides de Nivel desde SQLite (PrioridadManual) y Secuenciación Dinámica
        if results:
            op_ids = [r['IdOrden'] for r in results]
            scenario_id = request.GET.get('scenario_id')
            active_scenario = get_active_scenario(request, scenario_id=scenario_id)
            
            p_manual_db = PrioridadManual.objects.using('default').filter(
                scenario=active_scenario,
                id_orden__in=op_ids
            ).values('id_orden', 'nivel_manual')
            
            op_to_nivel = {p['id_orden']: p['nivel_manual'] for p in p_manual_db if p['nivel_manual'] is not None}
            
            # Agrupamos por MSTNMBR (pieza/artículo madre) para auto-secuenciar las operaciones del proceso
            from collections import defaultdict
            groups = defaultdict(list)
            for r in results:
                mst = r.get('MSTNMBR') or 0
                groups[mst].append(r)
                
            for mst, group in groups.items():
                # Ordenamos las operaciones por IdOrden (hoja de ruta natural del ERP)
                group.sort(key=lambda x: int(x.get('IdOrden') or 0))
                for i, r in enumerate(group):
                    oid = r['IdOrden']

                    # Nivel_Planificacion: SIEMPRE el valor nativo del ERP (TMAN002).
                    erp_val = r.get('Nivel_Planificacion')
                    r['Nivel_Planificacion'] = int(float(erp_val)) if erp_val is not None and str(erp_val).strip() not in ('', '0') else 0

                    # prioridad_articulo: independiente, desde override manual si existe
                    if oid in op_to_nivel:
                        r['prioridad_articulo'] = int(op_to_nivel[oid])
                    else:
                        r['prioridad_articulo'] = None

        return JsonResponse({'processes': results})
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'processes': [], 'error': str(e)}, status=500)

@csrf_exempt
def api_confirm_selected_tasks(request):
    """
    Guarda las OPs seleccionadas en el modelo PlannedTask y
    sus datos asociados (máquina, nivel, prioridad) en PrioridadManual,
    para que aparezcan en la grilla de planificación.
    """
    try:
        body = json.loads(request.body)
        id_ordens = body.get('id_ordens', [])
        piece_priorities = body.get('piece_priorities', {})
        selected_ops_by_article = body.get('selected_ops_by_article', {})
        scenario_id = body.get('scenario_id')
        project_code = body.get('project_code')
        force = body.get('force', False)

        active_scenario = get_active_scenario(request, scenario_id=scenario_id)

        if not active_scenario:
            from .models import Scenario
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            active_scenario = Scenario.objects.using('default').create(
                nombre=f"Planificación Temporal ({timestamp})",
                es_principal=False
            )
            request.session['last_scenario_id'] = str(active_scenario.id)
            request.session.modified = True
            print(f"DEBUG SELECTOR: Creado nuevo Scenario ID: {active_scenario.id}")
        elif str(active_scenario.id) != str(request.session.get('last_scenario_id', '')):
            request.session['last_scenario_id'] = str(active_scenario.id)
            request.session.modified = True
            print(f"DEBUG SELECTOR: Sesión sincronizada con Scenario ID: {active_scenario.id}")

        # -- PASO 1: Registrar proyecto en el escenario (SIEMPRE) --
        if not project_code and id_ordens:
            from django.db import connections
            with connections['production'].cursor() as cursor:
                cursor.execute("SELECT TOP 1 Formula FROM Tman050 WHERE Idorden = %s", [id_ordens[0]])
                row = cursor.fetchone()
                if row and row[0]:
                    project_code = str(row[0]).strip()

        if project_code and active_scenario:
            current_p = active_scenario.proyectos or ""
            p_list = [p.strip() for p in current_p.split(",") if p.strip()]
            if project_code not in p_list:
                p_list.append(project_code)
                active_scenario.proyectos = ",".join(p_list)
                active_scenario.save(using='default')
                print(f"DEBUG SELECTOR: Proyecto {project_code} registrado en escenario {active_scenario.id}")
                
            # PROTECCIÓN DE ESTADO: Asignar max_prioridad + 1 solo si es un proyecto nuevo
            from django.db.models import Max
            from .models import ProyectoPrioridad
            max_prio_proj = ProyectoPrioridad.objects.using('default').filter(
                scenario=active_scenario
            ).aggregate(Max('prioridad'))['prioridad__max'] or 0
            
            ProyectoPrioridad.objects.using('default').get_or_create(
                proyecto=project_code,
                scenario=active_scenario,
                defaults={'prioridad': max_prio_proj + 1}
            )

        # -- PASO 2: Obtener máquinas del ERP para los id_ordens --
        from django.db import connections
        op_maquina_map = {}
        if id_ordens:
            with connections['production'].cursor() as cursor:
                placeholders = ', '.join(['%s'] * len(id_ordens))
                cursor.execute(f"""
                    SELECT Idorden, Idmaquina
                    FROM Tman050
                    WHERE Idorden IN ({placeholders})
                """, id_ordens)
                for row in cursor.fetchall():
                    oid = str(row[0])
                    maq = str(row[1]).strip() if row[1] is not None else 'SIN ASIGNAR'
                    op_maquina_map[oid] = maq

        # -- PASO 3: Guardar cada OP con update_or_create --
        inserted_count = 0
        for oid in id_ordens:
            try:
                oid_str = str(oid)
                maquina = op_maquina_map.get(oid_str, 'SIN ASIGNAR')

                PlannedTask.objects.using('default').update_or_create(
                    id_orden=oid,
                    scenario=active_scenario,
                    defaults={'proyecto_code': project_code}
                )

                from django.db.models import Max
                max_art_prio = PrioridadManual.objects.using('default').filter(
                    scenario=active_scenario, maquina=maquina
                ).aggregate(Max('prioridad'))['prioridad__max'] or 0.0

                # PROTECCIÓN DE ESTADO: Usar get_or_create para no pisar prioridades manuales existentes
                PrioridadManual.objects.using('default').get_or_create(
                    id_orden=oid,
                    scenario=active_scenario,
                    maquina=maquina,
                    defaults={
                        'nivel_manual': 1,
                        'prioridad': max_art_prio + 100.0,
                        'orden_secuencia': 0,
                    }
                )

                inserted_count += 1
                print(f"DEBUG SELECTOR: OP {oid} guardada -> Máquina {maquina} | Proyecto {project_code}")

            except Exception as e:
                print(f"ERROR procesando OP {oid}: {e}")
                import traceback
                traceback.print_exc()
                continue

        # -- PASO 4: Aplicar prioridades de pieza (piece_priorities) --
        if piece_priorities and id_ordens:
            try:
                with connections['production'].cursor() as cursor:
                    pk_list = list(piece_priorities.keys())
                    placeholders = ', '.join(['%s'] * len(pk_list))
                    cursor.execute(f"""
                        SELECT MacroPK, MAX(Idorden)
                        FROM Tman050
                        WHERE MacroPK IN ({placeholders})
                        AND IsMacro = 1
                        GROUP BY MacroPK
                    """, pk_list)
                    master_ops = {}
                    for row in cursor.fetchall():
                        if row[1] is not None:
                            master_ops[str(row[0])] = str(row[1])

                for macro_pk, prio_val in piece_priorities.items():
                    if prio_val is None:
                        continue
                    ops = list(selected_ops_by_article.get(macro_pk, []))
                    master_op = master_ops.get(macro_pk)
                    if master_op and master_op not in ops:
                        ops.append(master_op)

                    for oid_str in ops:
                        try:
                            maquina = op_maquina_map.get(oid_str, 'SIN ASIGNAR')
                            
                            # PROTECCIÓN DE ESTADO: Usar update_or_create para permitir re-ordenamiento manual
                            PrioridadManual.objects.using('default').update_or_create(
                                id_orden=oid_str,
                                scenario=active_scenario,
                                maquina=maquina,
                                defaults={
                                    'nivel_manual': int(prio_val),
                                    'prioridad': float(prio_val),
                                    'orden_secuencia': 0,
                                }
                            )
                        except Exception as e:
                            print(f"ERROR prioridad OP {oid_str} (art {macro_pk}): {e}")
                            continue

            except Exception as e:
                print(f"ERROR consultando ERP para prioridades: {e}")
                import traceback
                traceback.print_exc()

        return JsonResponse({'status': 'ok', 'count': inserted_count, 'scenario_id': str(active_scenario.id)})

    except Exception as e:
        import traceback
        print(f"ERROR CONFIRMACIÓN: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)

def check_project_planning(request):
    """
    API to check if a project code already has planned tasks in a scenario.
    Permissive behavior: returns action='show' when project exists so the frontend
    navigates to display it, never blocking the interface.
    """
    project_code = request.GET.get('proyecto', '').strip()
    scenario_id = request.GET.get('scenario_id')
    
    if not project_code:
        return JsonResponse({'exists': False, 'action': 'select'})
        
    try:
        active_scenario = get_active_scenario(request, scenario_id=scenario_id)
        
        v1 = project_code
        v2 = project_code.replace('-', '.')
        v3 = project_code.replace('.', '-')
        codes = list({v1, v2, v3})
        
        exists = PlannedTask.objects.using('default').filter(
            proyecto_code__in=codes,
            scenario=active_scenario
        ).exists()
        
        print(f"DEBUG: check_project_planning - Proj: {project_code}, Codes: {codes}, Scenario: {active_scenario.nombre}, Exists: {exists}")
        
        # Contar cuántas OPs existen para este proyecto
        existing_count = 0
        existing_ids = []
        if exists:
            pt_qs = PlannedTask.objects.using('default').filter(
                proyecto_code__in=codes,
                scenario=active_scenario
            )
            existing_count = pt_qs.count()
            existing_ids = list(pt_qs.values_list('id_orden', flat=True))

        if exists:
            return JsonResponse({
                'exists': True,
                'action': 'show',
                'proyecto': project_code,
                'scenario': active_scenario.nombre,
                'scenario_id': str(active_scenario.id),
                'existing_ops_count': existing_count,
                'existing_ids': [str(eid) for eid in existing_ids]
            })
        else:
            return JsonResponse({
                'exists': False,
                'action': 'select',
                'proyecto': project_code,
                'scenario': active_scenario.nombre,
                'scenario_id': str(active_scenario.id) if active_scenario else None
            })
    except Exception as e:
        print(f"ERROR check_project_planning: {e}")
        return JsonResponse({'exists': False, 'action': 'select', 'error': str(e)})

@csrf_exempt
def api_clear_all_planning(request):
    """
    Clears ALL planned tasks and manual overrides for the active scenario.
    This effectively "empties" the planner for the current project(s) or scenario.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        scenario_id = body.get('scenario_id')
        active_scenario = get_active_scenario(request, scenario_id=scenario_id)
        
        from .models import PlannedTask, PrioridadManual, HiddenTask, TaskDependency, ProyectoPrioridad
        from django.db import transaction
        with transaction.atomic(using='default'):
            # 1. Delete all tasks from the planned list in this scenario
            PlannedTask.objects.using('default').filter(scenario=active_scenario).delete()
            
            # 2. Delete all manual overrides (machine moves, etc) for this scenario
            PrioridadManual.objects.using('default').filter(scenario=active_scenario).delete()

            # 3. Delete hidden tasks for this scenario
            HiddenTask.objects.using('default').filter(scenario=active_scenario).delete()

            # 4. Reset project priorities for this scenario
            ProyectoPrioridad.objects.using('default').filter(scenario=active_scenario).delete()

            # 5. Delete manual dependencies (These are currently global in the DB schema provided)
            # To be safe and meet the "reset" requirement, we clear them as they relate to the planning state.
            TaskDependency.objects.using('default').all().delete()
            
        return JsonResponse({'status': 'ok', 'message': 'Selección vaciada correctamente'})
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def api_delete_project_planning(request):
    """
    Deletes planned tasks, manual priorities, and hidden status for the specified projects in the active scenario.
    Removes the projects from the scenario's persistent list as well.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        body = json.loads(request.body)
        proyectos_raw = body.get('proyectos', '').strip()
        scenario_id = body.get('scenario_id')
        
        active_scenario = get_active_scenario(request, scenario_id=scenario_id)
        
        if not proyectos_raw:
            return JsonResponse({'error': 'No se especificaron proyectos para borrar.'}, status=400)
            
        proj_list = [p.strip() for p in proyectos_raw.split(',') if p.strip()]
        if not proj_list:
            return JsonResponse({'error': 'No se especificaron proyectos válidos.'}, status=400)
            
        # Collect all op ids to delete
        all_op_ids = set()
        codes_to_match = set()
        
        for proj in proj_list:
            v1 = proj
            v2 = proj.replace('-', '.')
            v3 = proj.replace('.', '-')
            for c in (v1, v2, v3):
                codes_to_match.add(c)
                
        # Query ERP to resolve Idorden
        ps = [f"%{c}%" for c in codes_to_match]
        if ps:
            from django.db import connections
            with connections['production'].cursor() as cursor:
                where_clauses = " OR ".join(["Formula LIKE %s"] * len(ps))
                sql = f"SELECT Idorden FROM Tman050 WHERE ({where_clauses})"
                cursor.execute(sql, ps)
                for row in cursor.fetchall():
                    all_op_ids.add(str(row[0]))
                    
        with transaction.atomic(using='default'):
            # Also find by matching proyecto_code in PlannedTask SQLite table directly to be extra thorough
            sqlite_ops = PlannedTask.objects.using('default').filter(
                scenario=active_scenario,
                proyecto_code__in=list(codes_to_match)
            ).values_list('id_orden', flat=True)
            
            for op in sqlite_ops:
                all_op_ids.add(str(op))
                
            # If we resolved some OP IDs, delete them from all three tables
            if all_op_ids:
                op_ids_list = list(all_op_ids)
                
                # Delete from PlannedTask
                PlannedTask.objects.using('default').filter(
                    scenario=active_scenario,
                    id_orden__in=op_ids_list
                ).delete()
                
                # Delete from PrioridadManual
                PrioridadManual.objects.using('default').filter(
                    scenario=active_scenario,
                    id_orden__in=op_ids_list
                ).delete()
                
                # Delete from HiddenTask
                HiddenTask.objects.using('default').filter(
                    scenario=active_scenario,
                    id_orden__in=op_ids_list
                ).delete()
                
            # Also clean up the projects text in the scenario model
            if active_scenario.proyectos:
                current_p = active_scenario.proyectos or ""
                p_list = [p.strip() for p in current_p.split(",") if p.strip()]
                # Remove deleted projects and their hyphen/dot variations
                new_p_list = [p for p in p_list if p not in proj_list and p.replace('-', '.') not in proj_list and p.replace('.', '-') not in proj_list]
                active_scenario.proyectos = ",".join(new_p_list)
                active_scenario.save(using='default')
                
        return JsonResponse({
            'status': 'ok',
            'message': f'Planificación de proyecto(s) {", ".join(proj_list)} borrada correctamente.',
            'deleted_ops_count': len(all_op_ids)
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def api_get_planned_projects(request):
    """
    Returns a list of distinct project codes that have tasks planned in the active scenario.
    """
    try:
        scenario_id = request.GET.get('scenario_id')
        active_scenario = get_active_scenario(request, scenario_id=scenario_id)
        
        # Get distinct projects from PlannedTask
        projects_qs = PlannedTask.objects.using('default').filter(
            scenario=active_scenario
        ).exclude(proyecto_code__isnull=True).exclude(proyecto_code='').values_list('proyecto_code', flat=True).distinct()
        
        # Get projects from scenario.proyectos field
        scenario_projects = []
        if active_scenario.proyectos:
            scenario_projects = [p.strip() for p in active_scenario.proyectos.split(',') if p.strip()]
            
        # Combine lists and ensure unique/clean results
        combined_set = set()
        for p in list(projects_qs) + scenario_projects:
            if p and p.strip():
                combined_set.add(p.strip())
                
        combined = sorted(list(combined_set))
        
        return JsonResponse({'projects': combined})
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def api_guardar_orden_manual(request):
    """
    API to save manual task order for a specific machine and scenario.
    Updates PrioridadManual.orden_secuencia and .prioridad fields.
    """
    print("--- PETICIÓN DE ORDEN RECIBIDA ---")
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        maquina_id = body.get('maquina_id')
        orden_ids = body.get('orden_ids', [])
        scenario_id = body.get('scenario_id')
        
        if not maquina_id or not orden_ids:
            return JsonResponse({'error': 'Missing required parameters'}, status=400)
            
        active_scenario = get_active_scenario(request, scenario_id=scenario_id)
        
        with transaction.atomic(using='default'):
            ordenes_actualizados = []
            for idx, id_orden in enumerate(orden_ids):
                # ... dentro del bucle for idx, id_orden in enumerate(orden_ids):
                try:
                    id_orden_int = int(id_orden)
                except ValueError:
                    print(f"No se pudo convertir {id_orden} a int, se omite")
                    continue
                
                # Calculate priority value (spread out to allow inserting later)
                prioridad = (idx + 1) * 1000.0
                
                # CAMBIO AQUÍ: Sumamos 1 para que el primer elemento sea 1 y no 0 (evita problemas de evaluación Falsy)
                orden_secuencia = idx + 1 
                
                print(f"Actualizando objeto ID {id_orden_int} con orden {orden_secuencia}")

                
                obj, created = PrioridadManual.objects.using('default').update_or_create(
                    id_orden=id_orden_int,
                    maquina=maquina_id,
                    scenario=active_scenario,
                    defaults={
                        'prioridad': prioridad,
                        'orden_secuencia': orden_secuencia
                    }
                )
                ordenes_actualizados.append({
                    'id_orden': id_orden_int,
                    'prioridad': prioridad,
                    'orden_secuencia': orden_secuencia
                })
        
        return JsonResponse({
            'status': 'success',
            'ordenes': ordenes_actualizados,
            'message': f'Orden manual guardado correctamente para {len(ordenes_actualizados)} tareas'
        })
    except Exception as e:
        import traceback
        print("Excepción en api_guardar_orden_manual:", traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def api_ordenar_automatico(request):
    """
    API to calculate and apply automatic task ordering based on priority rules.
    Returns the new ordered task IDs per machine for frontend to update.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        body = json.loads(request.body)
        scenario_id = body.get('scenario_id')
        active_scenario = get_active_scenario(request, scenario_id=scenario_id)
        
        # First, get all planned task IDs and their ERP data to get priorities
        planned_tasks_qs = PlannedTask.objects.using('default').filter(scenario=active_scenario)
        planned_ids = list(planned_tasks_qs.values_list('id_orden', flat=True))
        
        if not planned_ids:
            return JsonResponse({'status': 'ok', 'orden_por_maquina': {}})
            
        # Fetch ERP data for these tasks to get priority information
        from .services import get_planificacion_data
        erp_data = get_planificacion_data({'id_orden_in': planned_ids}, exclude_completed=False)
        
        # Group tasks by machine
        orden_por_maquina = {}
        
        # First, get the current machine assignments from PrioridadManual (to respect manual moves)
        maquina_assignments = {}
        for pm in PrioridadManual.objects.using('default').filter(
            scenario=active_scenario,
            id_orden__in=planned_ids
        ):
            maquina_assignments[pm.id_orden] = pm.maquina
        
        # Group ERP data by machine
        for task in erp_data:
            id_orden = task['Idorden']
            # Use machine from PrioridadManual if available, otherwise ERP's MAQUINA_ID
            maquina = maquina_assignments.get(id_orden, task.get('MAQUINA_ID', 'SIN ASIGNAR'))
            
            if maquina not in orden_por_maquina:
                orden_por_maquina[maquina] = []
            orden_por_maquina[maquina].append(task)
        
        # For each machine, sort tasks according to automatic rules:
        # Rule 1: First by Proyecto Prioridad (from ProyectoPrioridad)
        # Rule 2: Then by Nivel Planificación (from ERP data, lower first)
        # Rule 3: Then by Orden Visual (original ERP order)
        
        # First, get project priorities
        proyecto_prioridad_map = {}
        for pp in ProyectoPrioridad.objects.using('default').filter(scenario=active_scenario):
            proyecto_prioridad_map[pp.proyecto] = pp.prioridad
        
        orden_final_por_maquina = {}
        with transaction.atomic(using='default'):
            for maquina, tasks in orden_por_maquina.items():
                # Sort the tasks
                def get_sort_key(task):
                    # Project priority (lower number = higher priority, default to 999 if not set)
                    proyecto = task.get('ProyectoCode', '')
                    proj_prio = proyecto_prioridad_map.get(proyecto, 999)
                    # Nivel Planificación (lower number = higher priority, default to 999)
                    nivel = task.get('nivel_planificacion', task.get('Nivel_Planificacion', 999))
                    # Original order as tiebreaker
                    orden_visual = task.get('OrdenVisual', 0)
                    return (proj_prio, nivel, orden_visual)
                
                tasks_sorted = sorted(tasks, key=get_sort_key)
                
                # Extract IDs in order
                ids_ordenados = [task['Idorden'] for task in tasks_sorted]
                orden_final_por_maquina[maquina] = ids_ordenados
        
        # Eliminamos la escritura automática a la BD para no destruir el trabajo manual.
        # Si el usuario quiere guardar el orden automático, deberá presionar el botón de Guardar.
        
        return JsonResponse({
            'status': 'ok',
            'orden_por_maquina': orden_final_por_maquina
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


def estadisticas_produccion(request):
    """
    Dashboard de auditoría de tiempos ERP vs Realidades.
    Construye gráficos y sugerencias de ajuste para tiempos desviados consistentemente.
    """
    from .services import get_planificacion_data
    from collections import defaultdict

    # Fetch data but we might need past completed items too, or just the current plan
    # using get_planificacion_data
    try:
        data = get_planificacion_data({}, exclude_completed=False)
    except Exception as e:
        data = []

    from .models import PrioridadManual, PlannedTask
    # get_active_scenario is natively available in views.py
    active_scenario = get_active_scenario(request)
    
    planned_task_ids = list(PlannedTask.objects.using('default').filter(scenario=active_scenario).values_list('id_orden', flat=True))
    
    if not planned_task_ids:
        data = []
    else:
        try:
            data = get_planificacion_data({'id_orden_in': planned_task_ids}, exclude_completed=False)
        except Exception as e:
            data = []

    overrides = PrioridadManual.objects.using('default').filter(
        scenario=active_scenario, cantidad_producida_manual__isnull=False
    )
    manual_qty_map = {o.id_orden: o.cantidad_producida_manual for o in overrides}

    # Grouping structure
    machines_chart = defaultdict(lambda: defaultdict(lambda: {'std': 0.0, 'real': 0.0}))
    history = defaultdict(list)

    for item in data:
        t_id = int(item.get('Idorden', 0))
        qty_prod = manual_qty_map.get(t_id)
        
        if qty_prod is None:
            qty_prod = float(item.get('Cantidadpp') or 0.0)

        t_fichado_total = float(item.get('Total_Horas_Fichadas') or 0.0)
        t_std = float(item.get('Tiempo') or 0.0)
        
        # Lo graficamos siempre que tenga Tiempo STD o alguna Fichada, incluso si la cantidad prod es 0 aún
        # ya que la auditoría debe visualizar el plan vs la realidad.
        # Cantidad para la barra de standard = usa la cantidad lograda o si no 1 como proyeccion
        qty_used_for_std = qty_prod if qty_prod > 0 else float(item.get('cantidad_final') or 1.0)
        
        if t_std > 0 or t_fichado_total > 0:
            m_name = str(item.get('MAQUINAD', 'SIN ASIGNAR')).strip()
            articulo = str(item.get('Descri', '')).strip()
            # Combinar OP y Descri para evitar que se pisen en el grafico si son de misma descripcion
            proceso_label = f"OP {t_id} - {articulo}"
            
            std_time_total = t_std * qty_used_for_std
            
            machines_chart[m_name][proceso_label]['std'] += std_time_total
            machines_chart[m_name][proceso_label]['real'] += t_fichado_total
            
            if qty_prod > 0 and t_std > 0:
                t_real_unit = t_fichado_total / qty_prod
                desvio_pct = ((t_real_unit - t_std) / t_std) * 100.0
                history[(articulo, m_name)].append({'id': t_id, 'desvio': desvio_pct})

    # Prepare chart data (e.g. top 10 items per machine)
    chart_data_out = {}
    for m_name, items in machines_chart.items():
        labels = []
        std_data = []
        real_data = []
        for art, times in items.items():
            labels.append(art)
            std_data.append(round(times['std'], 2))
            real_data.append(round(times['real'], 2))
        
        chart_data_out[m_name] = {
            'labels': labels,
            'std': std_data,
            'real': real_data
        }

    # Generate suggestions
    sugerencias = []
    for k, ops in history.items():
        articulo, m_name = k
        ops_sorted = sorted(ops, key=lambda x: x['id'])
        
        if len(ops_sorted) >= 3:
            # Check last 3
            last_3 = ops_sorted[-3:]
            all_positive = all(o['desvio'] > 15.0 for o in last_3)
            all_negative = all(o['desvio'] < -15.0 for o in last_3)
            
            if all_positive:
                avg_desv = sum(o['desvio'] for o in last_3) / 3.0
                sugerencias.append(f"El proceso '{articulo}' en la máquina '{m_name}' tardó constantemente MÁS de lo previsto en las últimas 3 OPs. Se sugiere aumentar el tiempo estándar un {avg_desv:.1f}%.")
            elif all_negative:
                avg_desv = abs(sum(o['desvio'] for o in last_3) / 3.0)
                sugerencias.append(f"El estándar para '{articulo}' en '{m_name}' está holgado durante las últimas 3 OPs. Se sugiere reducir el tiempo estándar un {avg_desv:.1f}%.")

    return render(request, 'produccion/estadisticas.html', {
        'chart_data_json': json.dumps(chart_data_out),
        'sugerencias': sugerencias,
        'active_menu': 'estadisticas'
    })
